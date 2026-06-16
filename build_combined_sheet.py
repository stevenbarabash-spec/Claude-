import openpyxl
import pickle
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/all_rows.pkl', 'rb') as f:
    all_rows = pickle.load(f)

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
BLUE_BG, BLUE_TXT = "EAF4FB", "154360"
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT = Font(name="Calibri", size=10, color=BLUE_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Address", "Region / County",
    "City", "State", "ZIP", "Phone", "Owner Direct Email",
    "Website", "Business Type", "Years / Background",
    "✉ PERSONALIZATION HOOK", "📧 DRAFT EMAIL — Ready to Send",
    "Franchise?"
]
COL_W = [30, 28, 24, 22, 16, 6, 7, 18, 32, 28, 24, 36, 52, 80, 12]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Car Wash Leads"
ws.sheet_properties.tabColor = "117A65"

for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=NAVY_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 36

hook_col, email_col, state_col = 13, 14, 6

for i, row in enumerate(all_rows, start=2):
    state = row[5] or ""
    base_fill_color = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        if col_idx == hook_col:
            cell.fill = PatternFill("solid", fgColor=AMBER_BG)
            cell.font = HOOK_FONT
        elif col_idx == email_col:
            cell.fill = PatternFill("solid", fgColor=BLUE_BG)
            cell.font = EMAIL_FONT
        else:
            cell.fill = PatternFill("solid", fgColor=base_fill_color)
            cell.font = REGULAR_FONT
    ws.row_dimensions[i].height = 190

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(all_rows)+1}"

out_path = "/home/user/Claude-/All_CarWash_Leads_Combined.xlsx"
wb.save(out_path)
print(f"Saved {out_path} with {len(all_rows)} leads on one sheet.")
