"""
Multi-Industry M&A Leads Generator
Industries: HBC, Roofing, Plumbing, Car Wash, Manufacturing, Supermarkets, Healthcare Distribution, E-Commerce
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────────
NAVY_BG   = "1A2744"   # header background
WHITE_TXT = "FFFFFF"
AMBER_BG  = "FFF3CD"   # hook column
AMBER_TXT = "7D3C00"
BLUE_BG   = "EAF4FB"   # draft-email column
BLUE_TXT  = "154360"
ROW_ALT1  = "F7F9FC"
ROW_ALT2  = "FFFFFF"

HEADER_FONT  = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT    = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT   = Font(name="Calibri", size=10, color=BLUE_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Address", "Region / County",
    "City", "State", "ZIP", "Phone", "Owner Direct Email",
    "Website", "Business Type", "Years / Background",
    "✉ PERSONALIZATION HOOK", "📧 DRAFT EMAIL — Ready to Send",
    "Franchise?"
]

COL_W = [30, 24, 24, 22, 14, 6, 7, 16, 32, 28, 22, 34, 52, 80, 10]

# Tab colours per industry
TAB_COLORS = {
    "HBC":              "8E44AD",
    "Roofing":          "C0392B",
    "Plumbing":         "2471A3",
    "Car Wash":         "117A65",
    "Manufacturing":    "BA4A00",
    "Supermarkets":     "1E8449",
    "Healthcare Dist.": "1A5276",
    "E-Commerce":       "7D6608",
    "Summary":          "1A2744",
}

# ── Lead Data ─────────────────────────────────────────────────────────────────
# Each lead: [Business, Owner, Address, Region, City, State, ZIP, Phone,
#             Email, Website, Biz Type, Years/Background, Hook, DraftEmail, Franchise]

LEADS = {

# ─────────────────────────────────────────────────────────────────────────────
"HBC": [
    [
        "Independent Beauty Supply Co.",
        "Tyrone Rose",
        "93 Monticello Ave",
        "Hudson County",
        "Jersey City", "NJ", "07304",
        "(201) 333-3350",
        "tyronerose@icloud.com",
        "independentbeauty.com",
        "Health & Beauty Supply Distributor",
        "Since 1974 — 50+ years; one of the largest Black-owned & operated professional beauty distributors in the US; full-service salon product distribution; 2nd generation with Tyrone & Lisa Rose.",
        "52 years of independent beauty distribution — Tyrone built this into one of the top Black-owned HBC distributors in the country, right out of Jersey City.",
        (
            "Hi Tyrone,\n\n"
            "I came across Independent Beauty Supply and was genuinely impressed — 52 years as one of the country's premier Black-owned beauty distributors, all built independently out of Jersey City. That kind of longevity and reputation doesn't happen by accident.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the health and beauty distribution space across the tri-state area. I wanted to reach out to see if you might be interested in exploring a potential sale — now or at some point in the future.\n\n"
            "I also wanted to ask: are you currently open to expanding through acquisition? If there are competing distributors or complementary product lines in the region you've had your eye on, we may be able to help facilitate that as well.\n\n"
            "If you're open to a conversation, I'd love to set up a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Roofing": [
    [
        "J. Salvatore & Sons Roofing Co. Inc.",
        "Michael Salvatore",
        "1187 Yonkers Ave",
        "Westchester County",
        "Yonkers", "NY", "10704",
        "(914) 237-0683",
        "jsalvatoreroofing@gmail.com",
        "jsalvatoreroofing.com",
        "Family-Owned Residential & Commercial Roofing",
        "Since 1921 — over 100 years; 4th-generation Salvatore family; 35+ employees and 14 trucks; serves Westchester, NYC, and Connecticut.",
        "Salvatore & Sons has been roofing Westchester since 1921 — a 100-year family legacy now in its 4th generation, running 14 trucks out of Yonkers. That's rare.",
        (
            "Hi Michael,\n\n"
            "I was researching leading family-owned roofing operations in the Westchester area and Salvatore & Sons immediately stood out — over 100 years in business, four generations of Salvatores, 14 trucks on the road. That is an extraordinary legacy that very few businesses ever achieve.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established roofing businesses like yours in the New York area. I wanted to reach out to see if you might be open to exploring a potential sale — whether now or at some point down the road.\n\n"
            "I also wanted to ask: if growth through acquisition is something you've considered — taking over another operation, expanding territory — we may be able to help with that too.\n\n"
            "Either way, I'd love to connect for a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "A & J Reliable Inc.",
        "Andy Gallina",
        "160 Grandview Ave",
        "Rockland County",
        "Nanuet", "NY", "10954",
        "(845) 627-1500",
        "contact@ajreliable.com",
        "ajreliable.com",
        "Residential & Commercial Roofing (NY/NJ/CT)",
        "Since 1979 — 45+ years; Andy Gallina founded and still runs this tri-state roofing operation; serves New York, New Jersey, and Connecticut; residential, commercial, and industrial.",
        "45 years building a tri-state roofing operation from the ground up in Rockland County — Andy Gallina is the rare owner who's still at the helm four decades in.",
        (
            "Hi Andy,\n\n"
            "I was looking into top independently owned roofing contractors across the tri-state area and A & J Reliable came up repeatedly — 45 years in the business, still owner-operated, covering New York, New Jersey, and Connecticut. That's an impressive operation you've built.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire roofing businesses like yours in the region. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point in the future.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If there are companies in the area you've thought about absorbing, we may be able to help put something together.\n\n"
            "If you're open to a quick conversation, I'd love to set up a call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Tri-State Commercial Roofing Corp.",
        "Stuart Roberts",
        "143 Rt. 59, Bldg. 6C",
        "Rockland County",
        "Hillburn", "NY", "10931",
        "(845) 362-1730",
        "stuart@tri-stateroofing.com",
        "tri-stateroofing.com",
        "Commercial & Industrial Roofing (NY/NJ/CT)",
        "Founded by Stuart Roberts and wife Amy Roberts; 30+ years of commercial roofing across the Northeast; GAF-certified; known for long-term relationships with property managers and general contractors.",
        "Stuart and Amy Roberts built this together — a husband-and-wife team who turned a simple commitment to customer satisfaction into a 30-year commercial roofing empire spanning 3 states.",
        (
            "Hi Stuart,\n\n"
            "I was researching established commercial roofing operations in the tri-state area and Tri-State Commercial Roofing stood out — 30+ years in the business, built together with Amy, GAF-certified, and a reputation for exactly the kind of repeat client relationships that make a business genuinely valuable.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire commercial roofing companies like yours in the New York, New Jersey, and Connecticut market. I wanted to reach out to see if you and Amy might be open to exploring a potential sale — whether now or down the line.\n\n"
            "I also wanted to ask: if expansion through acquisition is something you've considered — absorbing a competitor, picking up a new service territory — we may be able to assist with that as well.\n\n"
            "If you're open to it, I'd love to set up a brief call at a time that works for you. Please let me know the best time and number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Vanguard Roofing (Vanguard Organization Inc.)",
        "Frank Algier",
        "1839 South Rd",
        "Dutchess County",
        "Wappingers Falls", "NY", "12590",
        "(845) 298-2926",
        "frank@vanguardroofing.com",
        "vanguardroofing.com",
        "Commercial Roofing (NY/NJ/CT/MA)",
        "Since 1972 — 50+ years of family ownership; serves New York, New Jersey, Connecticut, and Massachusetts; commercial, industrial, and institutional roofing.",
        "Frank Algier has been running Vanguard since 1972 — that's over 50 years of family ownership covering four states. A business this established and this durable is exactly what acquirers are looking for.",
        (
            "Hi Frank,\n\n"
            "I was researching long-standing family-owned roofing operations in the Northeast and Vanguard Roofing came up as one of the most enduring — in business since 1972, family-owned for over 50 years, covering New York, New Jersey, Connecticut, and Massachusetts. That's a remarkable legacy.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established commercial roofing operations like yours. I wanted to reach out to see if you might be open to exploring a potential sale — whether now or at some point in the future when the timing feels right.\n\n"
            "I also wanted to ask: are you interested at all in expanding through acquisition — perhaps picking up a regional competitor or adding capacity in a specific market? We can help with that too.\n\n"
            "I'd love to set up a quick call if you're open to it. Please let me know what time works best and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Plumbing": [
    [
        "Sanitary Plumbing & Heating Corp.",
        "Harris Clark",
        "571 Timpson Place",
        "Bronx County",
        "Bronx", "NY", "10455",
        "(212) 734-5000",
        "harris@sanitaryplumbing.com",
        "sanitary.nyc",
        "Plumbing & Heating (Commercial & Residential)",
        "Since 1929 — 3rd generation Clark family; Harris and brother Jonathan Clark operate as part of Omnia Mechanical Group (Sanitary Plumbing, Calray Boilers, Bolt Electric, Antler Motor & Pump); NYC Master Plumber.",
        "Harris is 3rd-generation running a 95-year-old Bronx plumbing institution — and Sanitary is just one piece of the Omnia Mechanical Group he's built across multiple trades.",
        (
            "Hi Harris,\n\n"
            "I came across Sanitary Plumbing and was genuinely impressed — 95 years in the Bronx, third generation of Clarks, and you've expanded it into something even bigger with the Omnia Mechanical Group. That kind of multi-trade platform is rare and incredibly well-positioned.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire plumbing and mechanical service businesses like yours in the New York market. I wanted to reach out to see if you might be open to exploring a potential sale of Sanitary or any part of the Omnia group — now or at some point in the future.\n\n"
            "I also wanted to ask: are you looking to continue expanding through acquisition? If there are other trades or companies in the city you've had your eye on, we may be able to help facilitate that.\n\n"
            "I'd love to connect for a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Harris Plumbing & Heating",
        "Pasquale Quaratino",
        "184 Glenmore Ave",
        "Kings County",
        "Brooklyn", "NY", "11212",
        "(718) 495-3400",
        "pasquale@harrisplumbingandheating.com",
        "harrisplumbingandheating.com",
        "Residential & Commercial Plumbing & Heating",
        "Since 1918 — over 100 years in Brooklyn; locally owned and operated; licensed NYC Master Plumber and fire suppression expert; serves Brooklyn, Manhattan, Queens, and the Bronx.",
        "Over 100 years in Brooklyn — Harris Plumbing has survived the Great Depression, two world wars, and the pandemic. Pasquale runs a century-old institution that buyers rarely get a shot at.",
        (
            "Hi Pasquale,\n\n"
            "I was researching long-established plumbing operations in New York and Harris Plumbing & Heating stood out immediately — over 100 years in Brooklyn, owner-operated, covering four boroughs. A business with that kind of history and reputation is truly one of a kind.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established plumbing companies like yours in the New York market. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point down the road.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If growing the footprint or adding service lines is something you've considered, we may be able to help make that happen.\n\n"
            "If you're open to a quick conversation, I'd love to set up a call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "JPI Plumbing & Heating Inc.",
        "Joseph Pietracatella",
        "4370A Victory Blvd",
        "Richmond County",
        "Staten Island", "NY", "10314",
        "(917) 509-6302",
        "joseph@jpiplumbing.com",
        "jpiplumbing.com",
        "Residential, Commercial & Industrial Plumbing",
        "Since 2012 — independently owned by Joseph Pietracatella; licensed NYC Master Plumber; serves Staten Island, NYC, and New Jersey; heating, cooling, and plumbing.",
        "Joseph Pietracatella built JPI from scratch since 2012 into a full-service plumbing, heating, and mechanical operation across Staten Island and NJ — the kind of owner-built business serious acquirers love.",
        (
            "Hi Joseph,\n\n"
            "I was looking into leading independently owned plumbing and heating companies on Staten Island and JPI came up as a strong, well-regarded operation — built from the ground up, serving residential and commercial clients across the island and into New Jersey.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire plumbing businesses like yours in the New York market. I wanted to reach out to see if you might be interested in exploring a potential sale — whether now or at some point in the future.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If taking over another plumbing company or adding capacity is something you've thought about, we may be able to help facilitate that.\n\n"
            "If you're open to a quick call, I'd love to connect at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Archer Plumbing & Heating Inc.",
        "Steven Salierno",
        "635 Old White Plains Rd",
        "Westchester County",
        "Tarrytown", "NY", "10591",
        "(914) 713-3040",
        "steven@archerpandh.com",
        "archerpandh.com",
        "Plumbing, Heating & Mechanical Services",
        "Since 2003 — independently owned by Steven Salierno; 20+ years serving Westchester County, New York City, and Connecticut; residential and commercial.",
        "Steven Salierno built Archer P&H from nothing in 2003 into a 20-year tri-area operation covering Westchester, NYC, and CT — that's the kind of owner-built story acquirers actively seek.",
        (
            "Hi Steven,\n\n"
            "I was researching established plumbing and mechanical contractors in Westchester and Archer Plumbing & Heating came up consistently — 20+ years, independently built by you since 2003, covering Westchester, New York City, and Connecticut. That's an impressive footprint for an owner-operated company.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire plumbing and mechanical service businesses like yours in the region. I wanted to reach out to see if you might be open to exploring a potential sale — now or down the line.\n\n"
            "I also wanted to ask: are you interested in growing through acquisition? If picking up another Westchester or Connecticut operation is something you've considered, we may be able to help with that.\n\n"
            "I'd love to set up a quick call at a time that works for you. Please let me know the best time and number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Caballero & Sons Plumbing & Heating Inc.",
        "Michael J. Caballero",
        "71 Bannard Street",
        "Monmouth County",
        "Freehold", "NJ", "07728",
        "(732) 294-7971",
        "office@caballeroplumbing.com",
        "caballeroplumbing.com",
        "Residential & Commercial Plumbing & Heating",
        "Since 1964 — 60+ years; founded by Arcelio & Evelyn Caballero as 'Bob Caballero Plumbing'; now run by Michael J. Caballero (President) and John P. Caballero (VP); Monmouth County institution.",
        "Arcelio and Evelyn Caballero started this in 1964 — 60 years later, Michael carries the family name forward in Monmouth County. A six-decade family business with this kind of roots is exactly what strategic buyers are hunting for.",
        (
            "Hi Michael,\n\n"
            "I was researching leading family-owned plumbing operations in Monmouth County and Caballero & Sons stood out immediately — 60 years in the business, founded by your family in 1964, and still independently owned and operated. That kind of longevity in the trades is rare.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established plumbing and heating companies like yours in the New Jersey market. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point in the future when the timing is right.\n\n"
            "I also wanted to ask: are you interested at all in expanding through acquisition? If absorbing another Monmouth County competitor or taking on additional service territory is something you've thought about, we may be able to help facilitate that.\n\n"
            "If you're open to a conversation, I'd love to set up a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Car Wash": [
    [
        "— See CarWash_Leads_WITH_EMAIL_DRAFTS.xlsx —",
        "Multiple Owners",
        "NY / NJ / CT",
        "Tri-State",
        "Various", "NY/NJ/CT", "—",
        "—",
        "— See master car wash file —",
        "—",
        "Independent Car Wash",
        "30 verified leads across NY, NJ, CT with custom emails. See CarWash_Leads_WITH_EMAIL_DRAFTS.xlsx and CT_Greenwich_CarWash_Leads.xlsx for full detail.",
        "30 car wash owner leads fully researched & emailed. Greenwich CT complete; Stamford, Darien, Norwalk, Westport, Bridgeport, New Haven, Waterbury, Hartford still in progress.",
        "See CarWash_Leads_WITH_EMAIL_DRAFTS.xlsx for all 30 fully drafted emails.",
        "No — Independent Only",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Manufacturing": [
    [
        "Worksman Cycles (Worksman Trading Corp.)",
        "Wayne Sosin",
        "8622 101st Ave",
        "Queens County",
        "Ozone Park", "NY", "11417",
        "(800) 962-2453",
        "cycles@worksman.com",
        "worksmancycles.com",
        "Industrial Bicycle & Tricycle Manufacturing",
        "Since 1898 — 125+ years; 4th-generation family-owned; America's oldest bicycle manufacturer; makes industrial cargo bikes, tricycles, and electric models for warehouses, factories, and institutions.",
        "America's oldest bicycle manufacturer, running since 1898 out of Queens — 125 years of family ownership. Wayne Sosin is stewarding a piece of American manufacturing history.",
        (
            "Hi Wayne,\n\n"
            "I came across Worksman Cycles while researching legacy American manufacturers and I had to reach out — 125+ years in business, 4th-generation family ownership, and still making industrial bikes and tricycles right here in Queens. That's one of the most remarkable manufacturing stories in New York.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established manufacturing businesses like yours in the region. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point in the future when the timing is right.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If adding production capacity, a complementary product line, or a regional facility is something you've considered, we may be able to help make that happen.\n\n"
            "If you're open to it, I'd love to set up a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
    [
        "Accurate Box Company Inc.",
        "Lisa Hirsh",
        "380 Chamberlain Ave",
        "Passaic County",
        "Paterson", "NJ", "07501",
        "(973) 278-5000",
        "lhirsh@accuratebox.com",
        "accuratebox.com",
        "Corrugated & Custom Packaging Manufacturing",
        "Since 1944 — 80+ years; 4th-generation family business; Lisa Hirsh (granddaughter of founder) serves as President & CEO since 1997; 400,000 sq ft Paterson facility; women-owned; produces high-graphics corrugated packaging.",
        "Lisa Hirsh took over from her grandfather and turned this 80-year-old Paterson box company into a 400,000 sq ft packaging powerhouse — 4th-generation, women-owned, and still privately held.",
        (
            "Hi Lisa,\n\n"
            "I was researching independently owned manufacturing operations in New Jersey and Accurate Box came up as a truly standout business — 80+ years, fourth generation, women-owned, 400,000 square feet in Paterson. What you've built on top of your grandfather's foundation is remarkable.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established manufacturing companies like yours in the region. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point in the future.\n\n"
            "I also wanted to ask: are you interested in growing through acquisition? If there are complementary packaging companies or regional competitors you've considered absorbing, we may be able to help facilitate that.\n\n"
            "If you're open to a conversation, I'd love to set up a quick call at a time that works for you. Please let me know the best time and number to reach you.\n\n"
            "Thank you, and looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent / Women-Owned",
    ],
    [
        "Innovation Foods",
        "4th-Generation Owner (Family-Held)",
        "4 Gorton Road",
        "Cumberland County",
        "Millville", "NJ", "08332",
        "(856) 455-2209",
        "info@innovation-foods.com",
        "innovation-foods.com",
        "Beverage Co-Manufacturing / Co-Packing",
        "Since ~1940s — 4th-generation family-owned; co-packs and co-manufactures for major brands including Oatly and GoodBelly; produces high-acid juices, waters, teas, and fermented beverages; $45M+ facility investment.",
        "A quiet 4th-generation family co-packer in South Jersey that now produces for Oatly and GoodBelly — the kind of behind-the-scenes manufacturer that strategic buyers target for its client contracts and production capacity.",
        (
            "Hello,\n\n"
            "I was researching beverage co-manufacturers in the Northeast and Innovation Foods is one of the most interesting businesses in the space — a fourth-generation family operation in Millville that now co-packs for household brands like Oatly and GoodBelly. That combination of family roots and marquee client relationships is rare.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established food and beverage manufacturing companies like yours in the region. I wanted to reach out to see if the family might be open to exploring a potential sale — now or at some point in the future.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If there are complementary production facilities or brands you've considered, we may be able to help make that happen.\n\n"
            "If you're open to a quick conversation, I'd love to set up a call at your convenience. Please let me know a good time and the best number to reach the right person.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent / Family-Held",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Supermarkets": [
    [
        "CTown Supermarkets — Individual Store Operators",
        "Various Independent Owners",
        "Multiple Locations",
        "NY Metro Area",
        "New York City", "NY", "—",
        "HQ: (914) 686-2400",
        "— Research individual store owners via CTown HQ —",
        "ctownsupermarkets.com",
        "Independent Supermarket Cooperative Member",
        "CTown operates 200 independently owned and operated stores; cooperative partner of Krasdale Foods; White Plains NY-based; individual store owners are the acquisition targets. Each store is independently owned — HQ can connect to store-level owners.",
        "CTown's 200 stores are each owned by independent operators — the cooperative model means each owner is a potential acquisition target. Direct outreach to HQ can unlock the right contacts.",
        (
            "Hello,\n\n"
            "I'm reaching out regarding the independently owned CTown locations operating in the New York area. I'm currently working with several qualified buyers who are actively looking to acquire independently operated supermarkets and grocery stores in the tri-state market.\n\n"
            "We're particularly interested in connecting with individual store owners who may be open to exploring a potential sale — now or in the future. We're also open to conversations with operators who are looking to acquire additional store locations and expand their footprint.\n\n"
            "If you could help connect us with individual CTown store operators, or if this is something that should be discussed directly at the store level, we'd welcome any guidance you can provide.\n\n"
            "I'd love to set up a brief call at your convenience. Please let me know a good time and the best number to reach the right person.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent Co-op Members",
    ],
    [
        "Key Food — Individual Store Operators",
        "Various Independent Owners",
        "Multiple Locations",
        "Tri-State Area",
        "Brooklyn / Queens / Bronx", "NY", "—",
        "HQ: (732) 906-5100",
        "— Research individual store owners via Key Food HQ —",
        "keyfood.com",
        "Independent Supermarket Cooperative Member",
        "Key Food is a retailer cooperative with 324 independently owned stores across NY, NJ, and CT; founded 1937 in Brooklyn; individual operator-owners are the acquisition targets; HQ based in Matawan, NJ.",
        "Key Food's 324 stores are each run by independent owners — a cooperative of individually held businesses dating to Brooklyn 1937. The opportunity is at the store level, not corporate.",
        (
            "Hello,\n\n"
            "I'm reaching out in connection with independently owned Key Food store operators across New York and New Jersey. I'm currently working with several qualified buyers who are actively seeking to acquire independently operated grocery stores in the tri-state market.\n\n"
            "We're particularly interested in connecting with individual Key Food store owners who may be open to exploring a sale — now or at some point down the road. We're also working with operators looking to grow through acquisition of additional store locations.\n\n"
            "We'd welcome any guidance on reaching individual store owners, or we're happy to discuss how to approach this at the store level if that's more appropriate.\n\n"
            "I'd love to set up a quick call at your convenience. Please let me know a good time and the best number to reach the right person.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent Co-op Members",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"Healthcare Dist.": [
    [
        "New England Medical Specialties Inc.",
        "Robert Proietto",
        "21 Commerce Drive, Suite 2",
        "New Haven County",
        "North Branford", "CT", "06471",
        "(203) 458-6094",
        "Pamilljr@yahoo.com",
        "newenglandmedicalspecialties.com",
        "Specialty Medical Sales & Distribution",
        "Since 1985 — 40 years; independently owned; specialty distributor of IV therapy, oncology, wound care, infection control, and fall prevention products; serves Connecticut hospitals and clinics; member of Alliance Medical Group.",
        "40 years quietly supplying Connecticut hospitals with specialty products — IV therapy, oncology, wound care — that most distributors don't carry. Robert Proietto runs an incredibly durable niche business.",
        (
            "Hi Robert,\n\n"
            "I was researching specialty healthcare distributors in Connecticut and New England Medical Specialties immediately stood out — 40 years in business, independently owned, serving CT hospitals with specialty product lines in IV therapy, oncology, and wound care that most distributors can't offer. That's a hard-to-replicate niche.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire specialty healthcare distribution businesses like yours in the Northeast. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point in the future.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If adding product lines, territories, or complementary distributors is something you've considered, we may be able to help facilitate that.\n\n"
            "If you're open to a conversation, I'd love to set up a quick call at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
],

# ─────────────────────────────────────────────────────────────────────────────
"E-Commerce": [
    [
        "— E-Commerce Leads: Research in Progress —",
        "Multiple Owners",
        "NY / NJ / CT",
        "Tri-State",
        "Various", "NY/NJ/CT", "—",
        "—",
        "— Requires targeted LinkedIn + Acquire.com research —",
        "acquire.com / dealstream.com",
        "Independent E-Commerce / Amazon FBA",
        "E-Commerce lead generation requires platform-specific research via Acquire.com, Empire Flippers, or LinkedIn targeting sellers with 7-figure Amazon FBA or D2C businesses in NY/NJ/CT. Most do not have a public web presence. Recommend targeted outreach via Acquire.com and LinkedIn Sales Navigator filters: Location = NY/NJ/CT, Industry = Retail/E-Commerce, Revenue = $1M–$10M.",
        "High-value e-commerce targets are typically found via Acquire.com, Empire Flippers, or LinkedIn Sales Navigator — not public web search. Owner emails are rarely public.",
        (
            "Hi [Owner Name],\n\n"
            "I came across your e-commerce business and was really impressed by what you've built — [X years] of independent growth in [product category] is no small feat.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire established e-commerce businesses like yours in the $1M–$10M revenue range. I wanted to reach out to see if you might be open to exploring a potential sale — now or at some point down the road.\n\n"
            "I also wanted to ask: are you interested in expanding through acquisition? If acquiring a complementary brand or product line is something you've considered, we may be able to help make that happen.\n\n"
            "If you're open to a quick call, I'd love to connect at your convenience. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you — looking forward to connecting.\n\n"
            "Best regards"
        ),
        "No — Independent",
    ],
],

}  # end LEADS

# ── Helper functions ──────────────────────────────────────────────────────────

def make_header_fill(color=NAVY_BG):
    return PatternFill("solid", fgColor=color)

def make_fill(color):
    return PatternFill("solid", fgColor=color)

def apply_header_row(ws, headers, col_widths, header_fill=None):
    if header_fill is None:
        header_fill = make_header_fill()
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 36

def apply_data_row(ws, row_idx, lead, col_count=15):
    alt_fill = make_fill(ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2)
    hook_col = 13
    email_col = 14

    for col_idx, value in enumerate(lead, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True,
                                   horizontal="left")
        if col_idx == hook_col:
            cell.fill = make_fill(AMBER_BG)
            cell.font = HOOK_FONT
        elif col_idx == email_col:
            cell.fill = make_fill(BLUE_BG)
            cell.font = EMAIL_FONT
        else:
            cell.fill = alt_fill
            cell.font = REGULAR_FONT

    ws.row_dimensions[row_idx].height = 200

def build_industry_sheet(wb, industry_name, leads, tab_color):
    ws = wb.create_sheet(title=industry_name)
    ws.sheet_properties.tabColor = tab_color

    apply_header_row(ws, HEADERS, COL_W)

    for i, lead in enumerate(leads, start=2):
        apply_data_row(ws, i, lead)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    return ws

def build_summary_sheet(wb):
    ws = wb.create_sheet(title="Summary & Strategy")
    ws.sheet_properties.tabColor = TAB_COLORS["Summary"]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55

    hdr_fill = make_header_fill()
    for col, val in enumerate(["Industry", "Leads Found", "Notes & Next Steps"], 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = HEADER_FONT
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 32

    summary_rows = [
        ("HBC (Health, Beauty & Cosmetics)", 1,
         "Independent Beauty Supply (Jersey City NJ) — Tyrone Rose — direct iCloud email confirmed. Expand to additional CT/NY HBC distributors."),
        ("Roofing", 4,
         "4 verified NY leads: Salvatore & Sons (Yonkers), A&J Reliable (Nanuet), Tri-State Commercial (Hillburn), Vanguard (Wappingers Falls). Add NJ/CT roofing next."),
        ("Plumbing", 5,
         "5 verified NY/NJ leads: Sanitary Plumbing (Bronx), Harris Plumbing (Brooklyn), JPI (Staten Island), Archer P&H (Tarrytown), Caballero & Sons (Freehold NJ). Harris Clark email confirmed."),
        ("Car Wash", 30,
         "30 fully verified leads in separate files: CarWash_Leads_WITH_EMAIL_DRAFTS.xlsx (NY/NJ/CT) + CT_Greenwich_CarWash_Leads.xlsx. Next CT city: Stamford."),
        ("Manufacturing", 3,
         "Worksman Cycles (Queens), Accurate Box (Paterson NJ), Innovation Foods (Millville NJ). All 4th-generation family-owned. Add CT manufacturers."),
        ("Supermarkets", 2,
         "CTown & Key Food cooperative summaries added. True leads require individual store-owner research via cooperative directories. Each store is independently owned."),
        ("Healthcare Dist.", 1,
         "New England Medical Specialties (North Branford CT) — Robert Proietto — 40 years, specialty distributor. Add NY/NJ healthcare dist. companies."),
        ("E-Commerce", 0,
         "Requires LinkedIn Sales Navigator or Acquire.com targeting. Filter: Location=NY/NJ/CT, Revenue=$1M-$10M, Industry=Retail/E-Commerce. Public emails not readily available."),
    ]

    for i, (ind, cnt, note) in enumerate(summary_rows, start=2):
        alt = make_fill(ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
        for col, val in enumerate([ind, cnt, note], 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = alt
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = THIN_BORDER
        ws.row_dimensions[i].height = 58

    return ws

# ── Build Workbook ────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for industry, leads in LEADS.items():
        tab_color = TAB_COLORS.get(industry, "1A2744")
        build_industry_sheet(wb, industry, leads, tab_color)

    build_summary_sheet(wb)

    out_path = "/home/user/Claude-/Multi_Industry_Leads_Master.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Tabs: {[ws.title for ws in wb.worksheets]}")
    total = sum(len(v) for v in LEADS.values())
    print(f"Total lead rows across all industries: {total}")

if __name__ == "__main__":
    main()
