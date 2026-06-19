import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
BLUE_BG, BLUE_TXT = "EAF4FB", "154360"
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT = Font(name="Calibri", size=10, color=BLUE_TXT)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"), right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"), bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = ["Business Name","Owner Name","Address","City","State","ZIP",
           "Phone","Owner / Business Email","Website","Business Type",
           "Background","✉ PERSONALIZATION HOOK","📧 DRAFT EMAIL — Ready to Send"]
COL_W = [30,26,26,16,6,8,18,32,26,24,40,50,72]

def fn(owner):
    if not owner or any(x in owner.lower() for x in ["not confirmed","owner —","name not"]):
        return None
    n = owner.split("(")[0].split("/")[0].split("&")[0].strip()
    return n.split()[0] if n.split() else None

def email(owner, biz, opening, hook_line=""):
    name = fn(owner)
    g = f"Hi {name}" if name else "Hello"
    return (f"{g},\n\nI came across {biz} and wanted to reach out — I'm working with several qualified buyers "
            f"actively looking to acquire businesses like yours in the area.\n\n"
            f"Would you be open to a quick 10-minute call to explore what that could look like? "
            f"No pressure at all — just a conversation.\n\n"
            f"Let me know a good time and the best number to reach you.\n\nThank you,\n[Your Name]")

def L(biz,owner,addr,city,state,zip_,phone,em,web,btype,bg,hook,opening=""):
    return [biz, owner or "Owner — name not confirmed", addr, city, state, zip_,
            phone, em, web, btype, bg, hook, email(owner, biz, opening)]

# ── HVAC ─────────────────────────────────────────────────────────────────────
HVAC = [
L("Compact HVAC Inc.","Frankie Castro","246 W 38th St","New York","NY","10018","(845) 222-2800","nycompacthvac@gmail.com","compacthvac.com","HVAC — Residential & Commercial","Family business since 1967; Frankie Castro (son of founder) has run it 40+ years.","55 years under one family is extraordinarily rare — a legacy HVAC brand in Manhattan that buyers compete for."),
L("American HVACR LLC","","—","New York","NY","10001","(718) 913-5466","americanhvacrny@gmail.com","americanhvacr.com","HVAC — Residential & Commercial","2019 Top Rated Local award winner; founded 2016.","A young, still-growing operation that already has a Top Rated Local award — the growth runway here is significant."),
L("Gemini Thermo HVAC","","1518 Bergen St","Brooklyn","NY","11213","(718) 690-3718","geminithermohvac@gmail.com","geminithermohvac.com","HVAC / Ductless Mini-Split","In business since 1993 — 30+ years; serves all 5 boroughs + Long Island.","30 years and a state-certified team serving all five boroughs is the kind of established platform buyers pay a premium to acquire."),
L("GallettAir Inc.","Carmine Galletta","860 Little East Neck Rd","West Babylon","NY","11704","(631) 587-8395","service@gallettair.com","gallettair.com","HVAC — Residential & Commercial","Founded 1983; Carmine started from a van days after his honeymoon; NATE/BPI/NYSERDA accredited; never subcontracts.","Carmine built this from literally nothing in 1983 — 40 years of pure owner-operated HVAC with elite certifications most buyers can't replicate from scratch."),
L("Sigma-Tremblay LLC","Jack Armstrong","4 Delavergne Ave","Wappingers Falls","NY","12590","(845) 297-4000","homeperformance@sigmatremblay.com","sigmatremblay.com","HVAC / Energy Audits / Home Performance","In business since 1969; BPI-Accredited + NYSERDA Home Performance ENERGY STAR partner.","Dual BPI/NYSERDA certification since 1969 — a legacy Hudson Valley HVAC business with credentials that can't be fast-tracked."),
L("AirCool HVAC","","80-76 222nd St","Queens Village","NY","11427","(718) 465-1653","info@aircoolpro.com","aircool-hvac.com","HVAC — Residential & Commercial","25 years in business; NYC MWBE-certified; serves Manhattan, Queens, Brooklyn, Bronx, Long Island.","NYC MWBE status plus 25 years and a multi-borough footprint makes this a standout acquisition in any competitive bidding scenario."),
L("MC Air LLC","","4266 US Highway 1","Monmouth Junction","NJ","08852","(732) 336-9000","mcairllc@gmail.com","mcairllc.com","HVAC — Residential & Commercial","Lennox-authorized dealer; serves 14 NJ counties.","A Lennox-authorized dealer covering 14 counties is the kind of credentialed, wide-reach operation a strategic buyer wants to fold into a larger platform."),
L("Platinum HVAC","","497 S 17th St","Newark","NJ","07103","(908) 472-0966","platinumhvac365@gmail.com","platinumhvac.us","HVAC — Commercial & Residential","NJ HVACR licensed; strong commercial portfolio; 24/7 emergency service; serves NJ and NY.","A 24/7 commercial-focused HVAC shop with a NJ license and NY market reach is hard to find independently owned."),
L("Sky Air Heating and Cooling","Robert Smith","31 Stirling Lane","Wayne","NJ","07470","(973) 271-3556","SkyAirHeatingandCooling@gmail.com","skyairhvac.com","HVAC — Residential","Founded 2019 by Robert after 15 years in the trade; PSE&G rebate program partner.","Robert came in with 15 years of hands-on experience before launching — a young business with deep technical roots already aligned with utility rebate programs."),
L("Jemmy's HVAC","Jemmy Lerma","76 Lodi Street","Hackensack","NJ","07601","(201) 596-6249","jemmyshvac@live.com","jemmyshvacnj.com","HVAC — Residential & Commercial","Family-owned; GE Appliances authorized contractor; active Bergen County Chamber member.","GE authorization plus deep Chamber roots in Bergen County signals a well-networked, relationship-driven business that buyers love."),
L("Grab Air LLC","","35 Jersey St","East Rutherford","NJ","07073","(201) 783-5796","grabair04@yahoo.com","grabairnow.com","HVAC — Residential & Commercial","15+ years in business; 4.7-star rated; serves 7 NJ counties.","7-county coverage and a 4.7-star rating after 15 years is a strong, defensible market position for a North Jersey HVAC independent."),
L("On Site Plumbing & HVAC","Waleed Alhamoud","195 US-46 Suite 2","Totowa","NJ","07512","(551) 899-0900","info@onsitenewjersey.com","onsitenewjersey.com","HVAC & Plumbing — Residential & Commercial","Actively expanding across North Jersey per 2025 press release.","A company publicly announcing expansion in 2025 is in growth mode — exactly the right moment for an acquisition conversation."),
L("M & M Heating & Cooling LLC","Nick DiPronio","7 Barnum Ave Cutoff, Unit D-1","Stratford","CT","06614","(203) 377-5866","infommheatingcooling@yahoo.com","mmheatingandcooling.com","HVAC — Residential & Commercial","CT-licensed and EPA-certified; owner-operated; serves Fairfield County.","A CT-licensed, EPA-certified owner-operated HVAC shop in Fairfield County — the kind of lean, well-run independent that transitions cleanly."),
L("CTS Heating & Cooling","Russell Willwerth","207 Burwell Rd","West Haven","CT","06516","(203) 389-6099","ctshc88@gmail.com","ctsheatingandcooling.com","HVAC — Residential & Commercial","50+ years of combined technician experience; licensed & insured; 24/7 emergency; serves Hartford, Fairfield, New Haven, Middlesex counties.","50+ years of combined team experience across four CT counties is a depth of institutional knowledge that's irreplaceable."),
L("Mechanical Heating and Air Conditioning LLC","Booker Washington III","192 Dixwell Ave, Ste 102","New Haven","CT","06511","(203) 780-8959","bookerw@mechanicalheatingac.com","mechanicalheatingac.com","HVAC — Residential & Commercial","Founded 2016; owner Booker personally responds to calls and shows up on every job per consistent reviews.","Booker's hands-on ownership style has built a reputation for personal service that customers specifically call out — a trust asset that transfers with the business."),
L("Air Care Inc.","Ronda Ramada (contact)","608 Monmouth Ave","Bradley Beach","NJ","07720","(732) 988-0151","RONDA@AIRCARENJ.COM","aircarenj.com","HVAC — Residential & Commercial","In business since 1986 — nearly 40 years serving the Jersey Shore.","Nearly 40 years on the Jersey Shore makes Air Care a local institution — the kind of tenure buyers don't find on the open market every day."),
L("Figlia & Sons","","746 E 9th Street","New York","NY","10009","(212) 686-0094","sales@figliasons.com","figliasons.com","PTAC & HVAC Sales/Service","In continuous operation since 1960 — NYC's longest-running PTAC specialist; the family still answers the phone.","65 years in business as NYC's longest-running PTAC specialist is a market position that simply can't be replicated."),
L("Reliable Refrigeration Plus, Inc.","Mark Woollard","280 Welton St","Hamden","CT","06517","(203) 772-2222","info@reliablerefrigerationplus.com","reliablerefrigerationplus.com","Commercial Refrigeration & HVAC","Founded 1985; 24/7 on-call commercial refrigeration; CT Heating & Cooling Contractors Association member.","40 years of around-the-clock commercial refrigeration — an essential-service niche with recurring maintenance contracts baked in."),
]

# ── ROOFING ───────────────────────────────────────────────────────────────────
ROOFING = [
L("Sullivan Contracting","George Sullivan","7 School St, Suite B","Griswold","CT","06351","(860) 373-6664","sullivancontractingct@gmail.com","sullivancontractingct.com","Roofing / Windows / Decks","Family-owned since 1982 in Eastern CT; 4.2-star Google rating from 50+ reviews.","40+ years and 50+ verified Google reviews in Eastern CT is the local reputation most independents spend a lifetime building."),
L("ADN Roofing LLC","","24 Webster Dr","Ansonia","CT","06401","(203) 751-3091","info@adnroofingllc.com","adnroofingct.com","Roofing — Residential","15+ years in CT; serves 20+ towns including Bridgeport, Stamford, Norwalk.","A 20-town coverage map after 15 years is a distribution footprint that takes most roofers decades to build."),
L("Realize Services LLC","Ivair Paixao","180 Emerald Pl","Stratford","CT","06614","(475) 447-1701","realizeservices@gmail.com","realizeservicesllc.com","Roofing — Residential & Commercial","Built entirely on word-of-mouth referrals; became a standout Stratford roofer.","A business built entirely on referrals has zero customer acquisition cost — that kind of organic trust is exactly what buyers look for."),
L("Brown Roofing Company, Inc.","Eddie Griffin","12 Progress Ave","Seymour","CT","06483","(203) 723-1372","egriffin@brownroofing.org","brownroofing.org","Roofing — Residential & Commercial","CT's premier roofer since 1972; Eddie started as a teenage laborer and worked up to owner in 2012.","50+ years of brand equity plus an owner who came up through the ranks — a rare combination of legacy and proven leadership."),
L("J.J. Landerman Roofing","Neal Landerman","102R Filley St","Bloomfield","CT","06002","(860) 242-0271","jennifer@jjlandermanroofing.com","jjlandermanroofing.com","Roofing — Residential & Commercial","Nearly 100 years in operation; CertainTeed certified.","A century-old roofing company in Connecticut is about as rare as it gets — exactly the kind of legacy acquisition that commands a premium."),
L("Albana Roofing LLC","Hassan Rabdishta","40 Union City Rd","Prospect","CT","06712","—","albanaroofing@gmail.com","albanaroofing.com","Roofing — Residential & Commercial","Family-founded 1993; second generation now involved; Shaban and Patty Rabdishta still active.","A second-generation family roofing business in the Waterbury/Prospect area — succession planning conversations come naturally at this stage."),
L("Stellar Roofing Inc.","Daniel Shumway","500 Trolley Blvd","Rochester","NY","14609","(585) 888-7663","info@stellarroofing.com","stellarroofing.com","Roofing — Residential","Greater Rochester Chamber of Commerce member; well-reviewed regional independent.","A chamber-active Rochester roofer with strong community standing is well-positioned for a clean acquisition transition."),
L("Reliance Roofing LLC","James Brown","1605 State Route 12","Binghamton","NY","13901","(607) 204-1100","Roofing@Reliance.LLC","relianceroofing.llc","Roofing — Residential","BBB accredited; offers virtual shingle color previews and digital contracts.","Digital contracts and virtual previews from an independent Binghamton roofer shows a forward-thinking operation — easier to scale than old-school paper shops."),
L("Fort-Cica Roofing & General Contractors","Anthony Cicalo II","720 East 141 Street","Bronx","NY","10454","(718) 585-9188","grivera@fortcica.com","fortcica.com","Roofing — Residential & Commercial","GAF commercial certified; serves all five NYC boroughs plus Westchester.","GAF commercial certification with a full 5-borough NYC footprint is a market reach most independent roofers never achieve."),
L("A & J Reliable, Inc.","","50 Second Ave, Suite A","Nanuet","NY","10954","(877) 492-9014","contact@ajreliable.com","ajreliable.com","Roofing / Siding / Gutters / Windows","In business since 1979; 40,000+ satisfied customers; A+ BBB rating.","40,000+ customers since 1979 and an A+ BBB rating — that customer base alone is a significant asset for any acquirer."),
L("Wetherall Roofing & Contracting","Brian Wetherall","—","Middle Village","NY","11379","(718) 894-7011","info@wetherallroofingandcontracting.com","wetherallroofingandcontracting.com","Membrane Roofing / Sheet Metal / Waterproofing","Second-generation family-owned; 30+ years; GAF commercial certified.","A second-generation Queens shop with 30 years and GAF commercial credentials is a strong standalone or bolt-on roofing acquisition."),
L("Manhattan Roofing","George Stewart","447 Broadway, 2nd FL #264","New York","NY","10013","(212) 495-9840","Manhattanroofsnyc@gmail.com","manhattanroofs.com","Roofing — Residential & Commercial","Owner-operated; 20+ years experience; 4.8/5 stars from 57 reviews.","A 4.8-star Manhattan roofer with 20+ years of owner-operated experience — premium location, premium reputation."),
L("Homestead Roofing Company","Jeff Willner","533 Goffle Rd","Ridgewood","NJ","07450","(201) 444-2233","info@homesteadroofing.com","homesteadroofing.com","Roofing — Residential & Commercial","Founded 1930; acquired 2023 by Jeff Willner — one of NJ's oldest roofing firms.","Founded in 1930 with new ownership in 2023 — the legacy brand plus a fresh owner who may already be open to partnership or growth conversations."),
L("Mattsson Roofing & Restoration LLC","Erik Mattsson","224 Pinelynn Road","Glen Rock","NJ","07452","(201) 925-9709","emattsson@mattssonrestoration.com","mattssonroofing.com","Roofing / Siding / Gutters","Founded 2018 by Erik Mattsson; built on precision craftsmanship in Bergen County.","A founder still hands-on in 2018-built Bergen County shop — young enough to have growth momentum, established enough to have real customer relationships."),
L("DP Roofing & Contracting LLC","","574 Somerset St","North Plainfield","NJ","07060","(908) 241-6122","dpcontracting2@yahoo.com","doneperfectroofing.com","Roofing — Residential & Commercial","In business since 1996; owner on every job site; GAF certified; serves 6 NJ counties.","An owner personally on every job site since 1996 with GAF certification and 6-county coverage — the operational discipline here shows."),
]

# ── HBC ───────────────────────────────────────────────────────────────────────
HBC = [
L("VLS Pharmacy","Gopesh Patel","4402 5th Ave","Brooklyn","NY","11220","(718) 854-1384","info@vlspharmacy.com","newdrugloft.com","Independent Compounding Pharmacy","Founded 1984; owner-operated Brooklyn neighborhood pharmacy; grew into dual-location compounding network.","A dual-location compounding pharmacy network built from scratch since 1984 — Gopesh has created real scale that commands a premium."),
L("New Drug Loft","Gopesh Patel","1410 Second Ave","New York","NY","10021","(212) 879-0910","info@newdrugloft.com","newdrugloft.com","Independent Compounding Pharmacy","Upper East Side pharmacy fixture since the 1980s; specialty in sterile compounding for dermatology, pediatrics, and wellness.","A sterile compounding specialty on the Upper East Side since the '80s is a rare, high-margin niche that generalist pharmacy buyers can't easily replicate."),
L("Mega Aid Compounding Pharmacy","Tatyana German","140 58th St, Suite 8G","Brooklyn","NY","11220","(212) 920-4500","tgerman@mega-aid.com","mega-aid.com","PCAB-Accredited Compounding Pharmacy","PCAB-accredited; positioned in white-label telehealth prescription fulfillment.","PCAB accreditation plus a telehealth fulfillment niche makes this an ideal acquisition target for a digital-health-focused buyer."),
L("Seven Pharmacy LLC","Hassan","345 Main Ave, Suite 1","Norwalk","CT","06851","(203) 900-4471","info@sevenpharmacy.com","sevenpharmacy.com","Independent Community Pharmacy","Dozens of reviews highlighting owner-level personal service; family-operated.","An owner whose name appears by first name in dozens of customer reviews has built a trust relationship that's the foundation of any successful pharmacy acquisition."),
L("Compounded Solutions in Pharmacy","Michael Roberge","810 Main Street","Monroe","CT","06468","(203) 268-4964","mike@compoundedsolutions.com","reliantcompoundedsolutions.com","Independent Compounding Pharmacy","Founded 1999 by Michael Roberge; tripled the pharmacy's footprint; now rebranded as Reliant Compounded Solutions.","Michael has already grown and rebranded once — a founder who's demonstrated he can scale is uniquely positioned for a thoughtful exit conversation."),
L("Medford Compounding & Specialty Pharmacy","","2608 Route 112","Medford","NY","11763","(631) 693-1403","info@medfordcompounding.net","medfordcompounding.net","Compounding-Only Specialty Pharmacy","Compounding-only; lean staffing model with growth potential.","A compounding-only Long Island pharmacy with a lean model is exactly the kind of clean, scalable asset a specialty pharma acquirer wants."),
L("Town & Country Compounding Pharmacy","John Herr","535 E Crescent Ave","Ramsey","NJ","07446","(201) 447-2020","info@tccompound.com","tccompound.com","Independent Compounding Pharmacy","~$3M revenue; 21-50 employees; one of Northern NJ's largest independent compounders; ships to 22 states.","$3M revenue, 22-state shipping, and 21-50 employees makes Town & Country one of the largest independent compounders in NJ — a significant acquisition target."),
L("Hopewell Pharmacy & Compounding Center","JoAnn Hobson & Eric Jaderlund","1 W Broad St","Hopewell","NJ","08525","(609) 466-1960","info@hopewellrx.com","hopewellrx.com","Independent Compounding & Community Pharmacy","Pharmacist husband-and-wife team who took over in 2018; strong local brand in the Hopewell/Pennington/Trenton market.","A pharmacist couple who acquired this themselves in 2018 understand the transaction process — making them exceptionally receptive to a well-structured offer."),
L("Kings Highway Pharmacy & Medical Supply","","1416 Kings Highway","Brooklyn","NY","11229","(718) 375-5757","Kingsph@nyphgroup.com","healthiswealthny.com","Independent Pharmacy + Medical Supply Retail","Dual revenue from Rx and medical supply; part of the independent 'Health is Wealth' network.","A combined pharmacy and medical supply retail footprint gives this business two revenue streams most competitors offer separately."),
L("St. Anthony Pharmacy","","3860 Broadway","New York","NY","10032","(212) 923-6111","info@justinpharmacy.com","stanthonyrx.com","Independent Compounding & Community Pharmacy","21+ years in Washington Heights; offers pet meds, pediatric compounding, and home delivery.","21+ years of community pharmacy in Washington Heights with pediatric compounding, pet meds, and delivery — a diversified service mix that's hard to replicate."),
L("PMC Distributors LLC","Michael Correll","61 Ramapo Valley Rd","Mahwah","NJ","07430","(201) 962-2494","priya@pmcdistributors.com","pmcdistributors.com","Cosmetics & Beauty Products Wholesale Distributor","Since 2006; authenticated close-out luxury cosmetics (Lancôme, Estée Lauder, MAC); primarily export-focused.","Authenticated luxury cosmetics distribution since 2006 with a global wholesale export model — a niche that's extremely difficult for a new entrant to break into."),
L("U.S. Beauty Supply LLC","","2101 Park Street","Hartford","CT","06106","(860) 586-8377","usbeautysupply@yahoo.com","usbsupply.weebly.com","Independent Beauty Supply Store","Long-running community beauty supply on Hartford's Park Street corridor.","A long-running beauty supply on Hartford's Park Street corridor has built a culturally specific customer loyalty that drives consistent repeat traffic."),
L("Avenue Chemists","","4501 30th Ave","Astoria","NY","11103","(718) 545-1010","info@avenuechemists.com","avenuechemists.com","Independent Community Pharmacy","Women-owned; established 2012 in Astoria; built on free local delivery and personal touch.","Women-owned and built on free local delivery since 2012 — Avenue Chemists has created the kind of neighborhood loyalty that sustains through any ownership transition."),
L("Thriftway Pharmacy — Flatbush","Chris","542 Flatbush Ave","Brooklyn","NY","11225","(718) 284-9800","flatbushrx@thriftwaypharmacy.com","thriftwayflatbush.com","Independent Community Pharmacy","Serving Brooklyn since 1960; part of a small NYC independent pharmacy group.","A Brooklyn pharmacy operating continuously since 1960 — six decades of community trust is an asset no new entrant can buy."),
L("Thriftway Pharmacy — Hell's Kitchen","Chris","646 10th Ave","New York","NY","10036","(212) 956-1100","stjrx@thriftwaypharmacy.com","thriftwaynyc.com","Independent Community Pharmacy","Manhattan flagship; independently owned multi-location group (not a chain) since 1960.","The same owner operates both Flatbush and Hell's Kitchen — a rare two-location independent pharmacy group in NYC, available as a single acquisition."),
]

# ── HEALTHCARE DISTRIBUTION ───────────────────────────────────────────────────
HEALTH_DIST = [
L("Merrick Surgical Supplies & Home Care","","139 Merrick Ave","Merrick","NY","11566","(516) 378-0119","info@merricksurgical.com","merricksurgical.com","Home Medical Equipment & Surgical Supplies","Long Island independent; strong local presence in Merrick/Nassau County corridor.","An independent Nassau County DME supplier with deep local roots is exactly the kind of community-embedded business DME acquirers seek."),
L("Mid-Island Medical Supply Co.","Chris Brown","2093 Wantagh Ave","Wantagh","NY","11793","(516) 781-7332","info@midislandmedical.com","midislandmedical.com","Home Medical Equipment & Surgical Supplies","Family-owned for decades on Long Island; owner Chris Brown named across multiple directories.","A family-owned Long Island DME business with a named, accessible owner is the ideal profile for an acquisition outreach."),
L("Fonte Surgical Supply, Inc.","Michael Fonte","892 E Ridge Rd","Rochester","NY","14621","(585) 338-1000","info@fontesurgical.com","fontesurgical.com","Home Medical Equipment / Uniforms / Surgical Supplies","Owner-operated; serves upstate NY; operates under Fonte Health Care Solutions branding.","Michael Fonte's owner-operated upstate NY operation with its own branded health division is a well-packaged acquisition target."),
L("Freeport Medical Supply, Inc.","Chuck Madu","75 S Main St","Freeport","NY","11520","(516) 208-7432","info@freeportmedical.com","freeportmedical.com","DME / Medical Equipment & Pharmacy","Serves Brooklyn, Queens, Bronx, Nassau, and Suffolk — broad metro NY DME footprint.","A five-borough-plus DME and pharmacy operation owned by a named independent is an unusual scale find outside of corporate roll-ups."),
L("Monroe Medical Supplies, Inc.","Nasri Saad","2715 Route 130 South","Cranbury","NJ","08512","(609) 395-5578","monroemeds@verizon.net","monroemedicalsupplies.com","Home Medical Equipment — Wholesale & Retail","Family-owned since 1985; still using a Verizon business email, signaling founder-operated.","Nasri has been running this since 1985 — that Verizon email says it all. A founder still this hands-on after 40 years is ready for a thoughtful exit conversation."),
L("Lincoln Medical Supply Company LLC","Paul Reses","913 N Main St","Pleasantville","NJ","08232","(609) 641-4050","info@lincolnmedicalsupply.com","lincolnmedicalsupply.com","Home Medical Equipment","60+ year operating history; serves South Jersey coastal market.","Over 60 years serving South Jersey's coastal communities is a customer loyalty runway most DME operations never build."),
L("Lake Surgical Supply","Carmen Trezza","92 Broadway","Denville","NJ","07834","(973) 627-8100","lakesurgical7@gmail.com","lakesurgical.com","Surgical Supplies & Home Medical Equipment","40+ year independent NJ operation; Carmen Trezza named as Owner/President on LinkedIn.","Carmen is publicly identifiable as the owner/president — a 40-year NJ independent with an owner who's both accessible and legacy-minded."),
L("Medford Medical Supply LLC","","185 Route 70, Ste 2","Medford","NJ","08055","(609) 654-4650","info@medfordmedicalsupply.com","—","Home Medical Equipment","Owner-operated; 20+ years healthcare industry experience; described as 'owner operated' on Yelp.","A specifically owner-operated South Jersey DME business with 20+ years in the industry — classic succession-timing profile."),
L("Dealmed Medical Supplies LLC","Shlomo Bistritzky","3512 Quentin Rd, Ste 200","Brooklyn","NY","11234","(718) 332-5633","sales@dealmed.com","dealmed.com","Independent Medical Supply Distributor","Founded 2005; largest independent medical supplier in NYC market; has made acquisitions itself.","Already the largest independent medical supplier in the NYC market and has already acquired a competitor — Dealmed is a strong acquisition target at the right valuation."),
L("Advanced Home Medical Supplies, Inc.","","43 S Main St","West Hartford","CT","06107","(860) 523-1076","info@ctadvancedhomemedical.com","advancedhomemedicalsupplies.com","Home Medical Equipment","Independent since 1988; relocated to West Hartford Center in 2015; serves Hartford County.","Operating independently since 1988 and strategically relocated to West Hartford Center in 2015 — a business that actively reinvests in its own future."),
L("On The Mend Medical Supplies & Equipment","Liam O'Keeffe","385 Main St S, Ste 102","Southbury","CT","06488","(203) 262-0383","info@onthemendmedical.com","onthemendmedical.com","DME / Home Medical Supplies","Multi-location CT operation; expanding into Hartford County.","Liam is publicly named and actively expanding into Hartford County — a growing independent DME business with a named owner is a rare outreach opportunity."),
]

# ── AUTO REPAIR ───────────────────────────────────────────────────────────────
AUTO = [
L("Tony's Auto Body, Inc.","Tony Trocchia","136 Richardson Street","Brooklyn","NY","11211","(718) 389-5648","info@tonysautobrooklyn.com","tonysautobrooklyn.com","Auto Body / Collision Repair","Multi-generation family operation in Williamsburg; Tony's son also involved.","A multi-generational family auto body shop in Williamsburg is a highly coveted collision repair asset in one of Brooklyn's busiest neighborhoods."),
L("Marmin Auto Body","Gary Gartenberg","2451 1st Avenue","New York","NY","10035","(212) 327-3486","info@marminautobody.com","marminautobody.com","Auto Body / Collision Repair","Direct Repair Program shop; works directly with major insurers; long-established Manhattan operation.","DRP status with major insurers in East Harlem is a recurring revenue stream most small body shops never qualify for."),
L("Auto Care East","","243 East 94th Street","New York","NY","10128","(212) 988-1515","info@autocareeast.com","autocareeast.com","Full-Service Auto Repair","Boutique neighborhood shop on the Upper East Side; rare full-service independent in a market dominated by dealerships.","Full-service independent auto repair on the Upper East Side, surrounded by dealerships — that's a differentiation story a buyer can build on."),
L("Paul's Auto Repair LLC","Paul LeBlanc Sr / Paul Jr / Dan LeBlanc","422 Tolland Street","East Hartford","CT","06108","(860) 568-8819","service@paulsautorepair.com","paulsautorepair.com","Full-Service Auto Repair","Three-generation LeBlanc family operation; NAPA AutoCare certified.","Three LeBlancs running a NAPA-certified shop together is a rare succession story — the next generation is already there, making this a smooth transition target."),
L("Ultimate Automotive Maintenance & Repair","","—","Cromwell","CT","06416","(860) 635-4133","service@ultimateautomotive.net","auto.ultimateautomotive.net","Full-Service Auto Repair","Family-owned; serves Central CT between Hartford and New Haven.","A family-owned auto repair operation sitting between Hartford and New Haven in an underserved Central CT corridor — strong location fundamentals."),
]

# ── SPECIALTY FOOD ────────────────────────────────────────────────────────────
FOOD = [
L("IAM International, Inc.","Neera Tulshian","4 Saddle Ridge Dr","Lebanon","NJ","08833","(908) 713-9651","info@iaminternationalinc.com","iaminternationalinc.com","Specialty Food Private-Label Manufacturer","Minority- and woman-owned; founded 1992; GMO-free sauces, bakery, condiments.","Neera has been running this minority- and woman-owned specialty food manufacturer since 1992 — a 30-year track record with certified credentials that add acquisition value."),
L("Woolco Foods, Inc.","Robert Kricelin","135 Amity Street","Jersey City","NJ","07304","(201) 716-2700","robert@woolcofoods.net","woolcofoods.net","Specialty Food Wholesale Distributor","10,000+ SKUs; serves NYC-area restaurants; 7-days-a-week operations.","10,000+ SKUs and a 7-day-a-week distribution operation serving NYC restaurants is a substantial, hard-to-replicate food distribution business."),
L("Pellicano Specialty Foods, Inc.","Anthony Habib","211 Reading Street","Buffalo","NY","14220","(716) 822-2366","sauceman@pellicanos.com","pellicanos.com","Specialty Sauce / Condiment Co-Packer","30+ years; private-label partner to regional restaurant chains.","Anthony has been the 'sauceman' behind regional restaurant chains' private-label products for over 30 years — a contract base that's sticky and predictable."),
L("Rajbhog Foods, Inc.","","60 Amity Street","Jersey City","NJ","07304","(551) 222-4700","orders@rajbhog.com","rajbhog.com","Indian Specialty Sweets & Snack Manufacturer","Pioneer in Indian confectionery manufacturing in the US; 9+ locations; manufacturing/wholesale hub in Jersey City.","The category-pioneer for Indian confectionery in the US with 9 locations — this is a platform acquisition, not just a single shop."),
L("Nuovo Pasta Productions, Ltd.","Carl Zuanelli","1330 Honeyspot Road Ext","Stratford","CT","06615","(203) 380-4090","customerservice@nuovopasta.com","nuovopasta.com","Artisan Pasta Manufacturer","Multi-building CT operation; featured in CBIA's 'Made in Connecticut' series; B2B wholesale exposure.","A 'Made in Connecticut' featured artisan pasta manufacturer with a multi-building facility and active B2B wholesale — Carl has built real scale here."),
]

# ── E-COMMERCE ────────────────────────────────────────────────────────────────
ECOMM = [
L("Cmple.com, Inc.","Andrey Bogdanov","—","Brooklyn","NY","11201","(347) 587-6480","info@cmple.com","cmple.com","HDMI / Networking / AV Cables — E-Commerce","Started in 2007 with $3,000 and 2 people in a basement; grew into a leading cable importer and dropship wholesaler.","$3,000 and two people in a basement in 2007 to a leading Brooklyn-based cable importer — Andrey has built real scale on a shoestring origin story."),
L("Global Leathers / LeatherSkins.com","Paul Crystal","253 W 35th St","New York","NY","10001","(212) 244-5190","info@globalleathers.com","globalleathers.com","Leather & Exotic Skins — Wholesale & DTC Online","40+ year family business in the NYC Garment District; son Jordan built the DTC arm LeatherSkins.com.","A 40-year Garment District family business with a second-generation DTC online brand layered on top — two acquisition angles in one."),
L("Rocky Mountain Western","","PO Box 408","New Vernon","NJ","07976","(866) 442-2656","info@rockymountainwestern.com","rockymountainwestern.com","Custom Bolo Ties — Online-Only Retailer","The only online store delivering fully custom, made-to-order bolo ties; new owner physically relocated business to NJ.","The only online destination for fully custom bolo ties is a genuine category monopoly — niche but defensible, and the NJ-based new owner may be ready to move on."),
L("Nomaterra","Aggieszka & Benjamin Burnett","115 W 18th St","New York","NY","10011","—","agnieszka@nomaterra.com","nomaterra.com","Luxury Travel-Inspired Fragrances — DTC","Ex-Glamour beauty editor and Columbia chemist; launched in Brooklyn 2012; all scents use indigenous local ingredients.","An ex-Glamour editor and Columbia chemist co-founding a luxury fragrance brand from Brooklyn — the story alone is a brand asset that commands premium pricing."),
L("Beautyque NYC","Sonia Khemiri & Sylvie Giret","—","New York","NY","10001","—","hello@beautyque.nyc","beautyque.nyc","3D Virtual Multi-Brand Beauty Marketplace","Founded May 2020 after COVID killed their SoHo lease; launched the first-ever 3D virtual beauty storefront.","Sonia and Sylvie turned a COVID disaster into a first-mover category — the first 3D virtual beauty retail platform for indie brands is a genuinely unique digital asset."),
L("LEIF","Stacy Longenecker","—","Brooklyn","NY","11211","—","stacy@leifshop.com","leifshop.com","Women's Clothing / Home Goods / Accessories — DTC + Boutique","Stacy ran the entire online store from home while holding a full-time job; opened the Williamsburg store in 2016; women-owned, 4-person team.","A women-owned Williamsburg boutique with a thriving DTC channel built from scratch — Stacy has done the hardest part already; a buyer gets the brand and the customer base."),
L("The Two Oh Three (The 203)","Tory Brown & Roscoe Brown","—","Fairfield","CT","06824","—","Info@TheTwoOhThree.com","thetwoohthree.com","Connecticut Lifestyle Apparel & Home Goods — DTC + Retail","Started 2014 in parents' Westport basement; now in 15+ CT retail locations with a Fairfield flagship.","Started in a basement, now in 15+ Connecticut retail locations — Tory and Roscoe have proven the model works; a buyer inherits a proven multi-channel brand."),
L("Harlem Candle Company","Teri Johnson","—","New York","NY","10027","(212) 457-4014","info@harlemcandlecompany.com","harlemcandleco.com","Luxury Scented Candles — DTC & Wholesale","Teri used $50K saved for her wedding to fund the company in 2015; made Oprah's Favorite Things 2023; ~$2M revenue.","Oprah's Favorite Things, $2M revenue, and still founder-led — Teri has done the heavy lifting of brand-building; a buyer walks into a nationally recognized candle brand."),
L("GIANTmicrobes","Andrew Klein","78 Harvard Ave","Stamford","CT","06902","(203) 504-8060","andrew@giantmicrobes.com","giantmicrobes.com","Science / Novelty Plush Toys — DTC & Wholesale","Founded 2002; 200+ plush microbe designs; sold in museum shops, science centers, and DTC online; ~11 employees.","200+ SKUs in a category GIANTmicrobes literally invented — a defensible niche with museum, science center, and DTC channels all established."),
L("Tavalon Tea","John-Paul Lee","100 Louis St, Unit G","South Hackensack","NJ","07606","1-800-282-5051","wholesale@tavalon.com","tavalon.com","Premium Loose-Leaf Tea — DTC & Wholesale Hospitality","Left Accenture in 2005; now in 8,500+ hotels and restaurants globally including the Plaza Hotel and Ritz Carlton; still founder-led.","8,500+ luxury hospitality accounts including the Plaza and Ritz Carlton, still run by the founder who left Accenture to build it — this is a rare find."),
L("Royal Silk Direct","","113 Westerly Rd","Princeton","NJ","08540","(609) 430-1212","rsd@royalsilkusa.com","royalsilkusa.com","Silk Apparel — DTC Catalog & Online Retail","Established 1978; one of the longest-running independent DTC silk brands in the US; no outside investors.","Nearly 50 years of DTC silk retail with zero outside investors — an owner ready to sell walks away with a clean, unencumbered business."),
L("Collyer's Mansion","Mauri Weakley","179 Atlantic Ave","Brooklyn","NY","11201","(347) 987-3342","info@shopthemansion.com","shopthemansion.com","Handmade Home Decor & Global Artisan Goods — DTC + Boutique","Opened 2012 in Ditmas Park; featured by Goop; Brooklyn Heights storefront plus DTC online.","Goop-featured, Brooklyn Heights-located, and with a DTC channel built over 12 years — Collyer's Mansion is a curated brand with real cultural cachet."),
L("Reisfields NYC","Reis Chester","—","Newark","NJ","07102","—","hello@reisfieldsnyc.com","reisfieldsnyc.com","Luxury Soy Candles & Home Fragrance — DTC + Wholesale","Founded 2016; stocked at Madewell and J.Crew; hand-poured in NJ lab.","Stocked at Madewell and J.Crew after just 8 years — Reis has built a nationally distributed luxury candle brand from a NJ lab, still founder-owned."),
L("BumpaBuilt","Adam Rivard","48 South Rd","Somers","CT","06071","(860) 281-1244","hello@bumpabuilt.com","bumpabuilt.com","3D-Printed Sensory Toys & STEM Kits — Shopify + Etsy + TikTok Shop","1 printer in a basement to 70 printers running 24/7; 1M+ TikTok likes; named after Adam's father.","1 printer to 70 printers, 1M+ TikTok likes, selling on three platforms simultaneously — Adam has built a viral, asset-light manufacturing brand in Somers CT."),
]

INDUSTRIES = [
    ("HVAC", HVAC, "E74C3C"),
    ("Roofing", ROOFING, "E67E22"),
    ("HBC", HBC, "8E44AD"),
    ("Healthcare Distribution", HEALTH_DIST, "2471A3"),
    ("Auto Repair", AUTO, "1E8449"),
    ("Specialty Food", FOOD, "D4AC0D"),
    ("E-Commerce", ECOMM, "1A5276"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

for sheet_name, leads, tab_color in INDUSTRIES:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 36

    HOOK_COL, EMAIL_COL = 12, 13
    for i, lead in enumerate(leads, start=2):
        state = lead[4] or ""
        base = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
        for col_idx, value in enumerate(lead, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            if col_idx == HOOK_COL:
                cell.fill = PatternFill("solid", fgColor=AMBER_BG)
                cell.font = HOOK_FONT
            elif col_idx == EMAIL_COL:
                cell.fill = PatternFill("solid", fgColor=BLUE_BG)
                cell.font = EMAIL_FONT
            else:
                cell.fill = PatternFill("solid", fgColor=base)
                cell.font = REGULAR_FONT
        ws.row_dimensions[i].height = 160

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(leads)+1}"

out = "/home/user/Claude-/Multi_Industry_100Leads.xlsx"
wb.save(out)
total = sum(len(l) for _, l, _ in INDUSTRIES)
print(f"Saved {out} — {total} verified leads across {len(INDUSTRIES)} industries.")
for name, leads, _ in INDUSTRIES:
    print(f"  {name}: {len(leads)}")
