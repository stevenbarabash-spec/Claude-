import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
BLUE_BG, BLUE_TXT = "EAF4FB", "154360"
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT = Font(name="Calibri", size=10, color=BLUE_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Address", "Region / County",
    "City", "State", "ZIP", "Phone", "Owner Direct Email",
    "Website", "Business Type", "Years / Background",
    "✉ PERSONALIZATION HOOK", "📧 DRAFT EMAIL — Ready to Send",
    "Franchise?"
]
COL_W = [30, 28, 24, 22, 16, 6, 7, 18, 32, 28, 24, 36, 52, 80, 12]

LEADS = [
    [
        "Pooran Car Wash",
        "Owner — call to confirm name (new ownership)",
        "760 Long Island Ave",
        "Suffolk County",
        "Deer Park", "NY", "11729",
        "(631) 392-4101",
        "Via poorancarwash.com/contact — call to obtain direct email",
        "poorancarwash.com",
        "Hand Car Wash & Auto Detailing",
        "Family-owned, recently changed hands to new ownership; full-service hand wash and detailing on Long Island.",
        "Pooran just changed hands to new ownership on Long Island — fresh owners are often the most open to a conversation about growth or exit options early on.",
        (
            "Hi there,\n\n"
            "I noticed Pooran Car Wash recently changed ownership here in Deer Park — congratulations on taking over a well-established Long Island operation.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
    [
        "Jiffy Car Wash",
        "Owner — call to confirm name (family-owned 70+ yrs)",
        "449 Elm St",
        "Fairfield County",
        "Stamford", "CT", "06902",
        "(203) 441-5875",
        "Via jiffycarwash.net — call to obtain direct email",
        "jiffycarwash.net",
        "Exterior Car Wash",
        "Family-owned and locally owned for over 70 years — one of the longest-running car washes in Stamford.",
        "70+ years in Stamford — Jiffy has outlasted generations of competitors and is exactly the kind of legacy operation buyers compete for.",
        (
            "Hello,\n\n"
            "I was researching long-standing car washes in Stamford and Jiffy Car Wash stood out immediately — over 70 years in business and still family-owned. That kind of longevity is rare and speaks volumes about the business you've built.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
    [
        "Apex Auto Spa",
        "Owner — call to confirm name (founded 2014)",
        "229 Greenwich Ave",
        "Fairfield County",
        "Stamford", "CT", "06902",
        "(203) 290-2024",
        "support@apexautospa.co",
        "apexautospa.co",
        "Car Wash & Detailing",
        "Family-owned, founded in 2014 by a group of car enthusiasts; trusted by 500+ local drivers in Stamford.",
        "Apex was built by car enthusiasts from the ground up in 2014 and has grown a loyal local following of 500+ regulars — a strong, well-loved local brand.",
        (
            "Hello,\n\n"
            "I came across Apex Auto Spa and was impressed by what you've built since 2014 — a loyal base of 500+ local drivers in Stamford is no small accomplishment for an independently owned shop.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
    [
        "Westwood Car Wash (Old Hook Car Wash)",
        "Nancy & Bernie Torraco",
        "148 Bergenline Ave",
        "Bergen County",
        "Westwood", "NJ", "07675",
        "(201) 664-0996",
        "Via westwoodcarwashnj.com/contact — call to obtain direct email",
        "westwoodcarwashnj.com",
        "Full-Service Car Wash",
        "Owned by Nancy and Bernie Torraco since 1976 — nearly 50 years of continuous family ownership in Bergen County.",
        "Nancy and Bernie have owned this since 1976 — almost 50 years in the same family. Businesses with this kind of single-family tenure are increasingly rare acquisition targets.",
        (
            "Hi Nancy and Bernie,\n\n"
            "I was researching long-standing, family-owned car washes in Bergen County and Westwood Car Wash stood out — nearly 50 years under your ownership since 1976. That kind of dedication and consistency is exactly what makes a business special.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
    [
        "Ho-Ho-Kus Hand Wash",
        "Owner — call to confirm name",
        "208 E. Franklin Tpke",
        "Bergen County",
        "Ho-Ho-Kus", "NJ", "07423",
        "(201) 857-0047",
        "info@hohokushandcarwash.com",
        "hohokushandcarwash.com",
        "Full-Service Hand Car Wash & Detail Center",
        "Independently owned full-service hand wash and detail center in Bergen County.",
        "A true full-service hand wash & detail center in one of Bergen County's most affluent zip codes — a premium location most acquirers would love to add to a portfolio.",
        (
            "Hello,\n\n"
            "I was researching independently owned hand car washes in Bergen County and Ho-Ho-Kus Hand Wash stood out for its full-service hand wash and detail offering in such a strong local market.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
    [
        "718 Super Wash",
        "Owner — call to confirm name (new ownership/management)",
        "138-77 Queens Blvd",
        "Queens County",
        "Queens", "NY", "11435",
        "(718) 523-2764",
        "Via superwash.com — call to obtain direct email",
        "superwash.com",
        "Hand Wash, Detailing, Ceramic Coating & Window Tinting",
        "Recently under new ownership and management; full-service offering including ceramic coating, vinyl wraps, and window tinting in Queens.",
        "718 Super Wash just changed hands and is already expanding services — ceramic coating, vinyl wraps, window tinting. New owners moving fast like this are often the most receptive to acquisition or growth conversations.",
        (
            "Hello,\n\n"
            "I noticed 718 Super Wash recently came under new ownership and you've already expanded into ceramic coating, vinyl wraps, and window tinting — that's an ambitious and smart move for a Queens car wash.\n\n"
            "I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
            "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
            "Thank you and looking forward to connecting."
        ),
        "No — Independent",
    ],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "New CarWash Leads (Batch 2)"
ws.sheet_properties.tabColor = "8E44AD"

for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=NAVY_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 36

hook_col, email_col, state_col = 13, 14, 6

for i, row in enumerate(LEADS, start=2):
    state = row[5] or ""
    base_fill_color = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        if col_idx == hook_col:
            cell.fill = PatternFill("solid", fgColor=AMBER_BG)
            cell.font = HOOK_FONT
        elif col_idx == email_col:
            cell.fill = PatternFill("solid", fgColor=BLUE_BG)
            cell.font = EMAIL_FONT
        else:
            cell.fill = PatternFill("solid", fgColor=base_fill_color)
            cell.font = REGULAR_FONT
    ws.row_dimensions[i].height = 190

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(LEADS)+1}"

out_path = "/home/user/Claude-/CarWash_Leads_Batch2_New.xlsx"
wb.save(out_path)
print(f"Saved {out_path} with {len(LEADS)} new leads.")
