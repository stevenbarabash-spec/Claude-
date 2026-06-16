import pickle
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/all_rows.pkl', 'rb') as f:
    all_rows = pickle.load(f)

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
RED_BG, RED_TXT = "FBE6E6", "922B21"   # FOMO follow-up email column (urgency red)
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
FOMO_FONT = Font(name="Calibri", size=10, color=RED_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Owner Direct Email", "City", "State",
    "Original Outreach Date", "Days Since Outreach", "Response Received?",
    "✉ PERSONALIZATION HOOK", "🔥 FOLLOW-UP EMAIL — FOMO (Ready to Send)"
]
COL_W = [30, 28, 32, 16, 6, 18, 14, 14, 50, 85]

def first_name(owner):
    if not owner:
        return "there"
    # strip parenthetical and titles, take first token before space/slash/&
    name = owner.split("(")[0].strip()
    name = name.split("/")[0].strip()
    name = name.split("&")[0].strip()
    first = name.split(" ")[0].strip()
    return first if first else "there"

def make_fomo_email(business, owner, hook):
    name = first_name(owner)
    return (
        f"Hi {name},\n\n"
        f"I wanted to follow up on my note from last week about {business}. I know things get busy, so I didn't want this to slip through the cracks.\n\n"
        f"Quick update: since I reached out, the buyer interest in independently owned car washes in your area has only gotten stronger — I'm now fielding inquiries from multiple qualified buyers actively looking to close on a deal in the next few months. A couple of them have specifically asked about businesses with exactly your profile ({hook.split('—')[0].strip() if '—' in hook else hook[:60]}).\n\n"
        f"I don't want you to miss the window if the timing ends up being right for you — buyer demand like this doesn't usually stay this high for long, and I'd rather you hear about the opportunity now than after it's already been placed elsewhere.\n\n"
        f"No pressure at all — even a 10-minute call would let you know where things stand and what your options might look like, whether that's selling now, down the road, or growing through an acquisition of your own.\n\n"
        f"Would you have a few minutes this week or next? Just let me know a good time and the best number to reach you.\n\n"
        f"Thank you, and I look forward to connecting."
    )

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FOMO Follow-Up Emails"
ws.sheet_properties.tabColor = "922B21"

for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=NAVY_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 36

hook_col, fomo_col = 9, 10
ORIGINAL_DATE = "2026-06-11"
DAYS_SINCE = 5

rows_out = []
for row in all_rows:
    business = row[0] or ""
    owner = row[1] or ""
    email = row[8] or ""
    city = row[4] or ""
    state = row[5] or ""
    hook = row[12] or ""
    fomo_email = make_fomo_email(business, owner, hook)
    rows_out.append([
        business, owner, email, city, state,
        ORIGINAL_DATE, DAYS_SINCE, "No response yet",
        hook, fomo_email
    ])

for i, row in enumerate(rows_out, start=2):
    state = row[4] or ""
    base_fill_color = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        if col_idx == hook_col:
            cell.fill = PatternFill("solid", fgColor=AMBER_BG)
            cell.font = HOOK_FONT
        elif col_idx == fomo_col:
            cell.fill = PatternFill("solid", fgColor=RED_BG)
            cell.font = FOMO_FONT
        else:
            cell.fill = PatternFill("solid", fgColor=base_fill_color)
            cell.font = REGULAR_FONT
    ws.row_dimensions[i].height = 170

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows_out)+1}"

out_path = "/home/user/Claude-/CarWash_Leads_FollowUp_FOMO.xlsx"
wb.save(out_path)
print(f"Saved {out_path} with {len(rows_out)} follow-up FOMO emails.")
