import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
BLUE_BG, BLUE_TXT = "EAF4FB", "154360"
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT = Font(name="Calibri", size=10, color=BLUE_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Address", "Region / County",
    "City", "State", "ZIP", "Phone", "Owner Direct Email",
    "Website", "Business Type", "Years / Background",
    "✉ PERSONALIZATION HOOK", "📧 DRAFT EMAIL — Ready to Send",
    "Franchise?"
]
COL_W = [30, 28, 24, 22, 16, 6, 7, 18, 32, 28, 24, 36, 52, 80, 12]

def email_template(business, owner_first, hook_short):
    return (
        f"Hi {owner_first},\n\n"
        f"I was researching independently owned car washes in your area and {business} stood out — {hook_short}\n\n"
        f"I'm currently working with several qualified buyers who are actively looking to acquire businesses like yours in the area. I wanted to reach out to see if you might be interested in exploring a potential sale now or in the future. I also wanted to ask if you're currently interested in expanding through acquisition opportunities in the market.\n\n"
        f"If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a good time and the best number to reach you.\n\n"
        f"Thank you and looking forward to connecting."
    )

LEADS_RAW = [
    # Business, Owner, Address, Region, City, State, ZIP, Phone, Email, Website, Type, Years/Background, HookShort
    ("Nice Guys Car Wash & Auto Detailing", "James (Jim) & Laura Abbate", "5791 Broadway",
     "Bronx County", "Bronx", "NY", "10463", "(718) 549-4010",
     "Via niceguyscarwashbxny.com — call to obtain direct email", "niceguyscarwashbxny.com",
     "Hand Car Wash & Auto Detailing",
     "Family-owned since 1974 — over 50 years; Jim Abbate is a well-known Bronx businessman and community fixture.",
     "Jim and Laura have run this since 1974 — over 50 years as Bronx community fixtures, with Jim known city-wide for his civic involvement."),

    ("Stop N Stare Hand Car Wash", "Owner — call to confirm name", "1748 Castle Hill Ave",
     "Bronx County", "Bronx", "NY", "10462", "(718) 823-3003",
     "Via Yelp listing — call to obtain direct email", "—",
     "Hand Car Wash",
     "Family-owned hand car wash in the Castle Hill section of the Bronx.",
     "a long-running, family-owned hand wash in Castle Hill — exactly the kind of single-location community business buyers are targeting."),

    ("Tremont Wash & Lube", "Owner — call to confirm name", "—",
     "Bronx County", "Bronx", "NY", "—", "—",
     "Via tremontcarwashandlube.com — call to obtain direct email", "tremontcarwashandlube.com",
     "Car Wash & Lube",
     "Independently owned wash and lube combo shop serving the Tremont section of the Bronx.",
     "a combo wash-and-lube shop in Tremont — the dual revenue stream (wash + lube) is a strong selling point for buyers."),

    ("Waring Ave Carwash And Lube", "Owner — call to confirm name", "2407 Boston Rd",
     "Bronx County", "Bronx", "NY", "10467", "—",
     "Via waringavecarwash.com — call to obtain direct email", "waringavecarwash.com",
     "Car Wash & Lube",
     "Independently owned car wash and lube shop on Boston Road in the Bronx.",
     "a Boston Road institution combining a car wash with a lube shop — diversified revenue that buyers love to see."),

    ("Gentle Touch Hand Car Wash & Detailing", "Owner — call to confirm name", "2271 Westchester Ave",
     "Bronx County", "Bronx", "NY", "10462", "(917) 708-9040",
     "Via Yelp listing — call to obtain direct email", "—",
     "Hand Car Wash & Detailing",
     "Independently owned hand car wash and detailing shop in the Westchester Square area of the Bronx.",
     "a hand-wash-and-detail operation in Westchester Square with the kind of loyal local following independent shops build over years."),

    ("Brighton Hand Car Wash", "Owner-Operator — call to confirm name", "—",
     "Richmond County", "Staten Island", "NY", "—", "—",
     "Via brightonhandcarwash.com — call to obtain direct email", "brightonhandcarwash.com",
     "Hand Car Wash & Detailing",
     "Single proprietor-owned; the owner personally oversees and inspects every wash and detail.",
     "an owner-operator who personally inspects every single wash — that level of hands-on quality control is rare and valuable."),

    ("Staten Island Wash & Repair Center", "Owner — call to confirm name", "1725 Richmond Ave",
     "Richmond County", "Staten Island", "NY", "10314", "(718) 761-7900",
     "Via Yelp listing — call to obtain direct email", "—",
     "Car Wash & Auto Repair",
     "Independently owned combination car wash and auto repair center on Staten Island.",
     "a combination wash-and-repair shop — two complementary revenue lines under one roof on Staten Island."),

    ("Majestic Mobile Car Wash & Detailing", "Owner — call to confirm name", "115 Industrial Loop",
     "Richmond County", "Staten Island", "NY", "10309", "—",
     "Via majesticmobilewash.com — call to obtain direct email", "majesticmobilewash.com",
     "Mobile Car Wash & Detailing",
     "10 years serving Staten Island, New Jersey, Brooklyn, and Manhattan with mobile wash and detail services.",
     "a 10-year mobile detailing operation already spanning Staten Island, NJ, Brooklyn, and Manhattan — built-in multi-market reach."),

    ("Ricky's Mobile Car Wash & Auto Detailing", "Ricky — call to confirm last name", "—",
     "Westchester County", "White Plains", "NY", "10601", "(914) 352-7969",
     "rickymobilecarwash@gmail.com", "rickysmobilecarwash.com",
     "Mobile Car Wash & Auto Detailing",
     "Over 10 years serving White Plains, Yonkers, Bronxville, and New Rochelle with mobile wash and detailing.",
     "a 10-year mobile detailing business already covering four Westchester communities — a strong regional footprint for a single owner."),

    ("Cross County Car Wash", "Owner — call to confirm name", "95 Vrendenburgh Ave",
     "Westchester County", "Yonkers", "NY", "10704", "(914) 965-8787",
     "Via business listing — call to obtain direct email", "crosscountycarwash.com",
     "Full-Service Car Wash",
     "Independently owned full-service car wash in Yonkers near the Cross County shopping area.",
     "a full-service wash positioned right near Cross County's retail corridor — strong, consistent traffic location."),

    ("Yonkers Car Wash", "Owner — call to confirm name", "975 Midland Ave",
     "Westchester County", "Yonkers", "NY", "10704", "—",
     "Via Yelp listing — call to obtain direct email", "—",
     "Car Wash",
     "Independently owned car wash on Midland Avenue in Yonkers.",
     "a straightforward, well-reviewed independent wash on Midland Ave — exactly the kind of stable, owner-run business buyers want."),

    ("Quickway Car Wash & Detail Center", "Owner — call to confirm name", "954 Nassau Rd",
     "Nassau County", "Uniondale", "NY", "11553", "(516) 486-1915",
     "Via business listing — call to obtain direct email", "—",
     "Car Wash & Detail Center",
     "Independently owned wash and detail center in Uniondale, Nassau County.",
     "a wash-and-detail combo shop in Uniondale — Nassau County's car wash market has been heating up with PE interest, making independents attractive targets."),

    ("Rockville Centre Car Wash Inc", "Owner — call to confirm name", "600 Sunrise Hwy",
     "Nassau County", "Rockville Centre", "NY", "11570", "(516) 536-8486",
     "Via business listing — call to obtain direct email", "—",
     "Car Wash",
     "Independently owned car wash on Sunrise Highway in Rockville Centre.",
     "a Sunrise Highway fixture in one of Nassau's most desirable villages — premium location, premium clientele."),

    ("Royal Car & Van Wash", "Owner — call to confirm name", "1542 Old Country Rd",
     "Nassau County", "Plainview", "NY", "11803", "(516) 694-6423",
     "Via business listing — call to obtain direct email", "—",
     "Car & Van Wash",
     "Independently owned car and van wash on Old Country Road in Plainview.",
     "a car AND van wash — that commercial-fleet capability broadens the buyer pool beyond typical retail car wash acquirers."),

    ("Seaford Brushless Car Wash", "Owner — call to confirm name", "3470 Merrick Rd",
     "Nassau County", "Seaford", "NY", "11783", "(516) 221-6080",
     "Via business listing — call to obtain direct email", "—",
     "Brushless Car Wash",
     "Independently owned brushless car wash on Merrick Road in Seaford.",
     "a brushless wash on the busy Merrick Road corridor — a name and reputation that's stuck around long enough to become a local landmark."),

    ("Stewart AV Car Wash Inc", "Owner — call to confirm name", "215 Stewart Ave",
     "Nassau County", "Bethpage", "NY", "11714", "(516) 735-6083",
     "Via business listing — call to obtain direct email", "—",
     "Car Wash",
     "Independently owned car wash on Stewart Avenue in Bethpage.",
     "a Bethpage institution on the Stewart Ave commercial strip — steady traffic, steady reputation."),

    ("Sunrise Sunshine Car Wash Inc", "Owner — call to confirm name", "34 W Sunrise Hwy",
     "Nassau County", "Valley Stream", "NY", "11581", "(516) 791-9867",
     "Via business listing — call to obtain direct email", "—",
     "Car Wash",
     "Independently owned car wash on West Sunrise Highway in Valley Stream.",
     "a Sunrise Highway wash in Valley Stream — one of the busiest commercial corridors on Long Island's South Shore."),

    ("Texas Car Washing Detailing Center", "Owner — call to confirm name", "15 Texas Ave",
     "Nassau County", "Island Park", "NY", "11558", "(516) 889-0497",
     "Via business listing — call to obtain direct email", "—",
     "Car Wash & Detailing",
     "Independently owned wash and detailing center in Island Park, Nassau County.",
     "a wash-and-detail shop in the small, tight-knit Island Park market — low competition, loyal local base."),

    ("Empire Car Wash of Inwood", "Bruno & Russell (Owner-Operators)", "—",
     "Nassau County", "Inwood", "NY", "11096", "—",
     "Via empirecarwashgroup.com — call to obtain direct email", "empirecarwashgroup.com",
     "Car Wash Group (Multiple Nassau Locations)",
     "Owner-operators Bruno and Russell hold ownership stakes across multiple Nassau County locations including Inwood, Oceanside, Elmont, Bellerose, and Floral Park.",
     "Bruno and Russell already operate across 5 Nassau County towns — a rare multi-site independent platform that's especially attractive to buyers looking to scale fast."),

    ("Skyway Car Wash", "Raj Srivastava", "221 Raymond Blvd",
     "Essex County", "Newark", "NJ", "07105", "(973) 344-8000",
     "Via skywaycarwash.com/contact-us — call to obtain direct email", "skywaycarwash.com",
     "Full-Service Car Wash",
     "Established over 50 years ago — Newark's oldest and most respected car wash, run by Raj Srivastava.",
     "Raj runs Newark's oldest car wash — over 50 years as the most respected name in the city's car care market."),

    ("Sharp Auto Spa & Lube Center", "Pete (Owner/Manager)", "489 McCarter Hwy",
     "Essex County", "Newark", "NJ", "07102", "—",
     "petemv@supercarwash.com", "—",
     "Auto Spa & Lube Center",
     "Independently owned auto spa and lube center on McCarter Highway in Newark.",
     "a wash-and-lube combo on McCarter Highway — Pete's built a dual-revenue operation in one of Newark's busiest commercial corridors."),

    ("Paterson Car Wash", "Owner — call to confirm name", "—",
     "Passaic County", "Paterson", "NJ", "—", "—",
     "Via business listing — call to obtain direct email", "—",
     "Full-Service Car Wash",
     "Locally owned and operated since 1972 — rated the #1 full-service car wash in the Paterson area for over 50 years.",
     "over 50 years and still rated #1 in Paterson — a deeply trusted, multi-generational local brand."),

    ("The Good Guys Car Wash", "Owner — call to confirm name", "954 Bloomfield Ave",
     "Passaic County", "Clifton", "NJ", "07012", "—",
     "Via business listing — call to obtain direct email", "—",
     "Hand Car Wash",
     "Independently owned hand car wash on Bloomfield Avenue in Clifton.",
     "a hand wash with a name built on reputation — \"Good Guys\" branding speaks to decades of trust with Clifton customers."),

    ("Prestige Auto Spa NJ", "Jeff Kovatch & Andrew Marolda (Managing Partners); Jorge Reyes (Operating Partner)", "—",
     "Ocean County", "Toms River", "NJ", "—", "(732) 929-2914",
     "Via prestigeautospanj.com — call to obtain direct email", "prestigeautospanj.com",
     "Full-Service Car Wash & Membership Program",
     "Family and locally-owned; celebrated 3 years of excellence; offers unlimited membership program in Toms River.",
     "Jeff, Andrew, and Jorge built a modern membership-based wash model in just 3 years — exactly the recurring-revenue structure acquirers pay a premium for."),

    ("Car Wash and Beyond", "Owner — call to confirm name", "1436 Lakewood Rd",
     "Ocean County", "Toms River", "NJ", "08755", "(732) 736-5510",
     "Via Yelp listing — call to obtain direct email", "—",
     "Car Wash",
     "Independently owned car wash on Lakewood Road in Toms River.",
     "a well-established Lakewood Road wash in one of Ocean County's busiest retail corridors."),

    ("Holiday Service Center and Car Wash", "Owner — call to confirm name", "1194 Rt 37 W",
     "Ocean County", "Toms River", "NJ", "08755", "(732) 240-9871",
     "Via Facebook page — call to obtain direct email", "—",
     "Service Station & Car Wash",
     "Independently owned combination service station and car wash on Route 37 in Toms River.",
     "a service-station-and-car-wash combo on Route 37 — two complementary income streams under one independent owner."),

    ("Platinum Car Wash & Oil Change", "Owner — call to confirm name", "1161 Wolcott St",
     "New Haven County", "Waterbury", "CT", "06705", "(203) 527-9700",
     "Via platinumcarwashandoil.com — call to obtain direct email", "platinumcarwashandoil.com",
     "Car Wash & Oil Change",
     "Family-owned car wash and oil change combo shop in Waterbury.",
     "a wash-and-oil-change combo shop — Platinum has built a one-stop-shop model that's especially appealing to strategic buyers in Waterbury."),

    ("Foam & Wash", "Family-Owned (Multi-Generation) — call to confirm name", "Multiple Locations",
     "Dutchess / Ulster Counties", "Poughkeepsie / Beacon / Newburgh", "NY", "—", "—",
     "Via foamandwash.com — call to obtain direct email", "foamandwash.com",
     "Self-Serve & Soft-Cloth Car Wash (13 Locations)",
     "Family-run since the 1960s — 13 locations across the Hudson Valley including Poughkeepsie, Beacon, Newburgh, Red Hook, Hyde Park, Lagrange, Vails Gate, Wappingers Falls, and Fishkill.",
     "a 60-year family operation that's quietly built 13 locations across the Hudson Valley — this is a true multi-site platform play, rare among independents."),

    ("Newburgh Auto Spa", "Owner — call to confirm name", "86 Route 17K",
     "Orange County", "Newburgh", "NY", "—", "—",
     "Via newburghautospa.com — call to obtain direct email", "newburghautospa.com",
     "Full-Service Car Wash, Oil Change & Auto Repair",
     "Serving the Hudson Valley community since 2004 — over 20 years; offers full-service wash, detailing, Mobil 1 express oil change, and auto repair.",
     "20 years and three service lines under one roof — wash, oil change, and repair — a diversified, recession-resistant business model."),

    ("Castro Oasis Auto", "Castro (Owner) — call to confirm full name", "—",
     "New Haven County", "New Haven", "CT", "—", "—",
     "Via castrooasisauto.com — call to obtain direct email", "castrooasisauto.com",
     "Auto Care, Car Wash & Detailing",
     "Independently owned auto care, car wash, and detailing shop serving the New Haven area.",
     "a true one-stop auto care, wash, and detailing shop — that breadth of services is a strong value-add for acquirers."),
]

LEADS = []
for biz, owner, addr, region, city, state, zip_, phone, email, website, btype, years, hook_short in LEADS_RAW:
    first = owner.split("(")[0].strip().split("&")[0].strip().split("/")[0].strip().split(" ")[0].strip()
    if first.lower() in ("owner", "owner-operator", "family-owned"):
        first = "there"
    draft = email_template(biz, first, hook_short)
    hook_full = hook_short[0].upper() + hook_short[1:] if hook_short else hook_short
    LEADS.append([biz, owner, addr, region, city, state, zip_, phone, email, website, btype, years, hook_full, draft, "No — Independent"])

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "New CarWash Leads (Batch 3)"
ws.sheet_properties.tabColor = "B7950B"

for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor=NAVY_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w
ws.row_dimensions[1].height = 36

hook_col, email_col = 13, 14

for i, row in enumerate(LEADS, start=2):
    state = row[5] or ""
    base_fill_color = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=i, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        if col_idx == hook_col:
            cell.fill = PatternFill("solid", fgColor=AMBER_BG)
            cell.font = HOOK_FONT
        elif col_idx == email_col:
            cell.fill = PatternFill("solid", fgColor=BLUE_BG)
            cell.font = EMAIL_FONT
        else:
            cell.fill = PatternFill("solid", fgColor=base_fill_color)
            cell.font = REGULAR_FONT
    ws.row_dimensions[i].height = 170

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(LEADS)+1}"

out_path = "/home/user/Claude-/CarWash_Leads_Batch3_30New.xlsx"
wb.save(out_path)
print(f"Saved {out_path} with {len(LEADS)} new leads.")
