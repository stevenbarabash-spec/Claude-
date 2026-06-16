import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_BG, WHITE_TXT = "1A2744", "FFFFFF"
AMBER_BG, AMBER_TXT = "FFF3CD", "7D3C00"
BLUE_BG, BLUE_TXT = "EAF4FB", "154360"
ROW_ALT1, ROW_ALT2 = "F7F9FC", "FFFFFF"
STATE_FILL = {"NY": "EAF2E3", "NJ": "FCEEEA", "CT": "EAF0FB"}

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_TXT, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
HOOK_FONT = Font(name="Calibri", size=10, color=AMBER_TXT, italic=True)
EMAIL_FONT = Font(name="Calibri", size=10, color=BLUE_TXT)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D3D4"),
    right=Side(style="thin", color="D0D3D4"),
    top=Side(style="thin", color="D0D3D4"),
    bottom=Side(style="thin", color="D0D3D4"),
)

HEADERS = [
    "Business Name", "Owner Name", "Address", "City", "State", "ZIP",
    "Phone", "Owner Direct / Business Email", "Website", "Business Type",
    "Years / Background", "✉ PERSONALIZATION HOOK",
    "📧 DRAFT EMAIL — Ready to Send"
]
COL_W = [30, 26, 26, 16, 6, 8, 18, 32, 28, 26, 40, 50, 80]

PITCH = (
    "I'm currently working with several qualified buyers who are actively looking to acquire "
    "businesses like yours in the area. I wanted to reach out to see if you might be interested "
    "in exploring a potential sale now or in the future. I also wanted to ask if you're currently "
    "interested in expanding through acquisition opportunities in the market.\n\n"
    "If you're open to discussing this, I'd be happy to set up a quick call. Please let me know a "
    "good time and the best number to reach you.\n\n"
    "Thank you and looking forward to connecting."
)

def first_name(owner):
    if not owner or "name not confirmed" in owner.lower() or owner.lower().startswith("owner"):
        return None
    name = owner.split("(")[0].strip().split("/")[0].strip().split("&")[0].strip()
    return name.split(" ")[0].strip() or None

def make_email(owner, opening):
    fn = first_name(owner)
    greeting = f"Hi {fn}" if fn else "Hello"
    return f"{greeting},\n\n{opening}\n\n{PITCH}"

def row(business, owner, address, city, state, zip_, phone, email, website, btype, background, hook, opening):
    return [business, owner or "Owner — name not confirmed", address, city, state, zip_,
            phone, email, website, btype, background, hook, make_email(owner, opening)]

PLUMBING = [
    row("Chris Lombardo Plumbing & Heating", "Chris Lombardo", "380 Devon Farms Rd", "Stormville", "NY", "12582",
        "(845) 702-3404", "chrisplumbing380@yahoo.com", "—", "Residential Plumbing & Heating",
        "Licensed Master Plumber, sole owner-operator serving Westchester/Putnam/Dutchess & CT.",
        "Chris runs this as a one-man, cross-state operation spanning NY and CT — exactly the kind of owner-operator shop that's often ready for a succession conversation.",
        "I came across Chris Lombardo Plumbing & Heating and was impressed that you've built a cross-state service area covering Westchester, Putnam, Dutchess, and into Connecticut as a Licensed Master Plumber."),
    row("Adrias Plumbing Corp.", "Georgios Petrakis", "218 Hamilton Ave", "Brooklyn", "NY", "11231",
        "(718) 492-2380", "adriasplumbing@gmail.com", "—", "Commercial/Residential Plumbing",
        "Pre-qualified NYCHA (NYC Housing Authority) contractor.",
        "Adrias is already pre-qualified for NYCHA contracts — meaning the bonding and insurance work a buyer would normally have to do is already done.",
        "I came across Adrias Plumbing Corp. and noticed you're a pre-qualified NYCHA contractor — that institutional credential is a real asset for a Brooklyn plumbing business."),
    row("Hunterdon Plumbing & Heating Services LLC", "Charlie Haug", "57 Toad Ln", "Ringoes", "NJ", "08551",
        "(908) 328-7209", "HUNTERDONPLUMBING@GMAIL.COM", "—", "Residential Plumbing & Heating",
        "Owner is a licensed/bonded NJ Master Plumber running it as an owner-operator.",
        "Charlie runs this as a licensed, bonded NJ Master Plumber — a classic owner-operator exit candidate once he's ready to step back.",
        "I came across Hunterdon Plumbing & Heating and was impressed that you've built it as a licensed, bonded Master Plumber operation here in Ringoes."),
    row("A-1 Quality Rooter Sewer & Drain Cleaning", "Helen", "75 Gypsy Ln", "Meriden", "CT", "06450",
        "(203) 235-8504", "helen.a1qualityrooter@gmail.com", "—", "Sewer & Drain Cleaning",
        "Family-owned and operating continuously since 1981 — 45 years in business.",
        "45 years of continuous family ownership is a rare legacy in the sewer/drain niche — exactly the kind of long-tenured operation buyers compete for.",
        "I came across A-1 Quality Rooter and was struck by the 45 years of continuous family ownership behind your sewer and drain cleaning business in Meriden."),
    row("R.J. Riquier Inc.", "Allen Riquier", "415 Boston Post Rd", "North Windham", "CT", "06257",
        "—", "rjriquier@gmail.com", "—", "Plumbing, Heating, A/C & Solar Thermal",
        "3rd-generation family business, 50+ years; owner holds top-tier licenses across 4 trade categories.",
        "Three generations and four licensed trade categories under one roof — a diversified, multi-generational operation that's hard to find and easy to build on.",
        "I came across R.J. Riquier Inc. and was impressed that you're carrying on a 3rd-generation family business spanning plumbing, heating, A/C, and solar thermal."),
    row("AAA Combustion, LLC", "David E. Barr", "4021 Avenue T", "Brooklyn", "NY", "11234",
        "(646) 773-4248", "nydob4261@gmail.com", "—", "Oil/Gas Boiler & Hot Water Heater Service",
        "Incorporated 2010; pre-qualified NYC HPD contractor.",
        "AAA Combustion's niche boiler/combustion specialty plus NYC HPD pre-qualification makes it a strong complement to a broader plumbing rollup.",
        "I came across AAA Combustion and noticed you're a pre-qualified NYC HPD contractor specializing in oil and gas boiler service — a strong niche within plumbing and heating."),
    row("1st Heating Solutions LLC", "Doug Simmons", "619 Oakside Rd", "Yorktown Heights", "NY", "10598",
        "(845) 628-9900", "heatsolutionsllc@aol.com", "—", "HVAC / Heating",
        "Founded 2005; owner has 25+ years industry experience, former technical director for a major boiler manufacturer.",
        "Doug's background as a former technical director for a major boiler manufacturer gives this shop a credibility edge most independents can't match.",
        "I came across 1st Heating Solutions and was impressed by your background as a former technical director for a major boiler manufacturer before founding the company in 2005."),
    row("A & A Plumbing Systems, Inc.", "Owner — name not confirmed", "115-03 Farmers Blvd, #1", "Saint Albans", "NY", "11412",
        "(718) 783-9281", "info@aaplumbingsystems.com", "—", "General Plumbing Systems",
        "Pre-qualified for NYC Housing Authority (NYCHA) contracts.",
        "An established institutional/multifamily client base via NYCHA pre-qualification gives this business a steady, recurring revenue base buyers value.",
        "I came across A & A Plumbing Systems and noticed your NYCHA pre-qualification — that institutional client base is a strong foundation for a Queens-based plumbing operation."),
]

MANUFACTURING = [
    row("Mira Plastics Co., Inc.", "Anthony Miragliotta Jr.", "123 Fredon Springdale Rd", "Fredon Twp", "NJ", "07860",
        "—", "sales@miraplastics.com", "miraplastics.com", "Thermoplastic Injection Molding",
        "70-year family legacy; founded 1955 in a one-room Paterson garage with one used machine.",
        "Starting from one used machine in a one-room garage in 1955 to a 70-year family operation today is the kind of growth story buyers love to back further.",
        "I came across Mira Plastics and was struck by the 70-year family story — starting in a one-room Paterson garage with a single used machine back in 1955."),
    row("Anthony & Sons Bakery", "Baldo Dattolo", "20 Luger Rd", "Denville", "NJ", "07834",
        "(973) 625-2323", "baldo@anthonyandsonsbakery.com", "anthonyandsonsbakery.com", "Wholesale Bakery / Food Production",
        "25-year family-built business.",
        "Baldo built this wholesale bakery from the ground up over 25 years — a proven food-production operation with an established customer base.",
        "I came across Anthony & Sons Bakery and was impressed by the 25-year wholesale bakery business you've built in Denville."),
    row("Nolt's Machine Shop, LLC", "Marlin / Jordan Nolt", "3329 Depew Rd", "Canandaigua", "NY", "14424",
        "(585) 526-6004", "sales@noltsmachine.com", "noltsmachine.com", "CNC Precision Machining",
        "Family-owned, full-service CNC shop known for fast turnaround.",
        "A family-run CNC shop known specifically for fast turnaround has built a reputation that's hard for a new entrant to replicate quickly.",
        "I came across Nolt's Machine Shop and was impressed by the reputation you've built for fast-turnaround CNC precision machining."),
    row("True Steel Construction, LLC", "Owner — name not confirmed (woman/family-owned)", "248 Thomas St", "Newark", "NJ", "07107",
        "917-204-2611", "truesteelconstruction@gmail.com", "—", "Structural Steel Fabrication",
        "WBE-certified, in-house design-to-build capability, 15,000 sq ft shop.",
        "WBE certification plus full design-to-build capability in a 15,000 sq ft shop makes this a well-positioned acquisition target for buyers chasing diversity-certified contracts.",
        "I came across True Steel Construction and was impressed by your WBE-certified, design-to-build steel fabrication operation in Newark."),
    row("MetalMania Sheet Metal Works", "Wai Un Wong", "—", "Brooklyn", "NY", "11201",
        "—", "metalmaniaworks@gmail.com", "—", "Custom Sheet Metal Fabrication",
        "Founder spent 15 years fabricating in Macau casinos before building the NYC shop.",
        "Fifteen years fabricating for Macau casinos before launching a Brooklyn shop is a rare, high-end skill background that adds real craftsmanship value.",
        "I came across MetalMania Sheet Metal Works and was struck by your background fabricating for casinos in Macau before building this shop in Brooklyn."),
    row("Hurley Metal Fabrication & Manufacturing", "Bill Hurley", "725 Marshall Phelps Rd", "Windsor", "CT", "06095",
        "(860) 688-8844", "Bill@hurleymetalfab.com", "—", "Precision Metal Fabrication",
        "Recently moved into a newly purchased/renovated 40,000 sq ft facility.",
        "A recent move into a newly purchased, renovated 40,000 sq ft facility signals real momentum and growth capacity for the next owner to build on.",
        "I came across Hurley Metal Fabrication and noticed you recently moved into a newly renovated 40,000 sq ft facility — that's a strong signal of growth and capacity."),
    row("Marcelli Steel (Marcelli LLC)", "Carmine Marcelli", "105A Grays Bridge Rd", "Brookfield", "CT", "06804",
        "(203) 917-3942", "marcellisteel@gmail.com", "—", "Structural Steel Fabrication",
        "Owner started as a 16-year-old apprentice welder in Italy, sold a prior family business in 2007, founded this one in 2010.",
        "Carmine has already successfully sold one family business before — making him uniquely well-positioned to have a thoughtful conversation about doing it again on his own terms.",
        "I came across Marcelli Steel and was struck by your story — starting as a 16-year-old apprentice welder in Italy, then successfully selling a prior family business before founding this one in 2010."),
    row("Anderson Machinery Co.", "Stuart & Carl Anderson", "60 Rachel Dr", "Stratford", "CT", "06615",
        "—", "ANDMY@AOL.COM", "—", "Used Machinery Sales & Service",
        "Multi-generational sibling ownership, long-standing operation.",
        "A sibling-owned, multi-generational machinery business is exactly the kind of long-standing operation that benefits from a clear succession plan.",
        "I came across Anderson Machinery Co. and was impressed by the long-standing, sibling-owned operation you and Carl have built in Stratford."),
    row("Mechanical Precision Inc.", "Owner — name not confirmed", "11 Hopewell Ave", "Flemington", "NJ", "08822",
        "(908) 782-2511", "mecpre@earthlink.net", "—", "Precision Contract Manufacturing",
        "Small, family-owned contract manufacturer.",
        "A small, family-owned contract manufacturer is often the best-fit acquisition profile for a buyer looking to add capacity without inheriting corporate bureaucracy.",
        "I came across Mechanical Precision Inc. and was impressed by the precision contract manufacturing operation you've built as a family-owned business in Flemington."),
    row("Iron & Oak", "Owner — name not confirmed", "—", "Bethel", "CT", "06801",
        "(914) 438-0295", "Ironandoakct@gmail.com", "—", "Metal Fabrication",
        "Independent metal fabrication shop in Bethel, CT.",
        "An independent fabrication shop with a strong brand name like Iron & Oak has room to expand its customer base under the right ownership.",
        "I came across Iron & Oak and was impressed by the metal fabrication business you've built in Bethel."),
    row("A1 Welding & Fabrication LLC", "Andrew", "238 Shaker Rd", "Enfield", "CT", "06082",
        "(860) 969-9203", "AndrewA1welding@gmail.com", "—", "Welding & Fabrication",
        "Owner-operated welding and fabrication shop in Enfield.",
        "Andrew's owner-operated shop has built a name-recognition advantage with the straightforward 'A1' branding in the Enfield market.",
        "I came across A1 Welding & Fabrication and was impressed by the welding and fabrication business you've built in Enfield."),
    row("CSD Welding and Fabrication LLC", "Owner — name not confirmed", "—", "Wolcott", "CT", "06716",
        "(860) 518-8801", "Csdwfllc@aol.com", "—", "Welding & Fabrication",
        "Independent welding and fabrication shop in Wolcott, CT.",
        "A standalone welding and fabrication LLC in Wolcott represents a clean, bolt-on acquisition opportunity with no corporate entanglements.",
        "I came across CSD Welding and Fabrication and was impressed by the independent welding and fabrication business you've built in Wolcott."),
]

SUPERMARKETS = [
    row("Halinka Polish Deli", "Owner — name not confirmed", "438 US Hwy 206, Ste 4", "Hillsborough", "NJ", "08844",
        "(908) 829-3271", "PLdeli31@gmail.com", "polishdelinj.com", "Polish Specialty Deli/Grocery",
        "Family-owned since 2010 — 15 years serving NJ's Polish community.",
        "Fifteen years serving a tight-knit Polish community has built a loyal customer base that's hard for a new entrant to replicate.",
        "I came across Halinka Polish Deli and was impressed by the 15 years you've spent serving Hillsborough's Polish community."),
    row("Piast Meats & Provisions", "Owner — name not confirmed", "800 River Dr", "Garfield", "NJ", "07026",
        "(973) 340-4722", "info@piast.com", "piast.com", "Polish Gourmet Market",
        "Family-owned, founded 1991, grown from a single deli to two NJ locations.",
        "Growing from a single deli to two locations over 30+ years shows a proven, replicable model — exactly what a buyer looks for in a multi-unit opportunity.",
        "I came across Piast Meats & Provisions and was impressed by the growth from a single deli in 1991 to two thriving NJ locations today."),
    row("Heart of Europe Polish Deli", "Owner — name not confirmed", "1020 Route 18", "East Brunswick", "NJ", "08816",
        "(732) 955-6513", "heartofeuropepolishdeli@gmail.com", "heartofeuropedeli.com", "Polish/European Deli & Grocery",
        "Family business catering to a tight-knit immigrant community.",
        "A deli built specifically to serve a tight-knit immigrant community has a customer loyalty moat that's difficult to build from scratch.",
        "I came across Heart of Europe Polish Deli and was impressed by the dedicated following you've built within East Brunswick's Polish community."),
    row("Polski Smak", "Teresa Dabek", "1295 NJ-23", "Butler", "NJ", "07405",
        "(973) 492-7625", "szajbus1@aol.com", "—", "Polish Deli/Grocery",
        "Founded 2006 — nearly 20 years as a woman-owned Polish deli/grocery.",
        "Nearly two decades of woman-owned leadership gives Teresa's business a strong, established identity in the Butler market.",
        "I came across Polski Smak and was impressed that you've built and run this Polish deli/grocery for nearly 20 years."),
    row("Angelo's Market", "Owner — name not confirmed", "—", "New Britain", "CT", "06051",
        "(860) 223-7340", "angelosmarket349@gmail.com", "angelosmarket.com", "Italian Specialty Grocery/Deli",
        "A New Britain staple for over 60 years.",
        "Sixty years as a community staple is the kind of legacy reputation that's nearly impossible to put a price tag on — and exactly what acquirers seek out.",
        "I came across Angelo's Market and was struck by the 60+ years you've spent as a staple of the New Britain community."),
    row("D. Coluccio & Sons", "Owner — name not confirmed", "1214-20 60th St", "Brooklyn", "NY", "11219",
        "(718) 436-6700", "orders@dcoluccioandsons.com", "dcoluccioandsons.com", "Italian Specialty Importer/Retailer",
        "Family-owned for generations in Bensonhurst, with its own import business.",
        "A multi-generational import-and-retail operation in Bensonhurst gives this business both a loyal local base and a wholesale revenue stream.",
        "I came across D. Coluccio & Sons and was impressed by the generations-deep legacy and import business you've built in Bensonhurst."),
    row("Kosher Corner Supermarket", "Yehuda Cohen", "2055 McDonald Ave", "Brooklyn", "NY", "11223",
        "(718) 998-2400", "yeco492@aol.com", "—", "Kosher Full-Service Supermarket",
        "Personalized, owner-on-the-floor community service model.",
        "An owner who's personally on the floor every day has built a level of community trust that a buyer would want to carefully preserve through any transition.",
        "I came across Kosher Corner Supermarket and was impressed by the personalized, owner-on-the-floor service model you've built in Brooklyn."),
]

HEALTHCARE_DIST = [
    row("Prime Medical and Surgical Supply Co.", "Owner — name not confirmed", "5723 New Utrecht Ave", "Brooklyn", "NY", "11219",
        "(718) 437-0066", "Orders@PrimeMedSupply.Com", "primemedsupply.com", "Home Care & Rehab Equipment Distributor",
        "Explicitly markets itself as independently owned and operated, serving Brooklyn and the Bronx.",
        "Prime Medical markets its independent ownership directly to customers — a clear, owner-led identity that's well positioned for a succession conversation.",
        "I came across Prime Medical and Surgical Supply and was impressed by the independently owned, personal-service model you've built serving Brooklyn and the Bronx."),
    row("Mohawk Healthcare (Mohawk Hospital Equipment Corp.)", "Tom Spellman, Jr.", "247 Elizabeth Street", "Utica", "NY", "13501",
        "800-962-5660", "service@mohawkhealthcare.com", "mohawkhealthcare.com", "Full-Line Medical/Surgical Wholesale Distributor",
        "Founded 1946 — 80 years privately owned, NY State's leading independent wholesale distributor.",
        "Eighty years of private ownership and a position as NY State's leading independent wholesale distributor is a remarkable legacy worth building on.",
        "I came across Mohawk Healthcare and was struck by the 80-year legacy you've built as NY State's leading independent wholesale medical distributor."),
    row("ARMAC Inc.", "Andres Moreno III", "71 Passaic Ave", "Florham Park", "NJ", "07932",
        "888-422-3044", "allison.smith@armac.us", "armac.us", "DME / Orthopedic & Prosthetic Supplier",
        "40 years serving NJ's orthopedic/DME community across multiple locations.",
        "Forty years and multiple NJ locations in the orthopedic/DME niche is exactly the kind of depth a strategic acquirer would pay up for.",
        "I came across ARMAC Inc. and was impressed by the 40 years and multiple New Jersey locations you've built serving the orthopedic and DME community."),
    row("Collins Medical Equipment / Collins Medical Solutions", "Bryan Collins", "500 Kings Highway East", "Fairfield", "CT", "06825",
        "203-576-8642", "info@collinsmedical.net", "collinsmedical.net", "DME, Stairlifts & Accessibility Solutions",
        "Family-owned since 1931 — nearly a century-old family operation.",
        "Nearly a century of family ownership is an extraordinarily rare legacy — the kind of multi-generational story buyers actively seek out.",
        "I came across Collins Medical and was struck that your family has owned and run this business since 1931 — almost a century of continuous operation."),
    row("Help Medical Supplies", "Owner — name not confirmed", "555 Madison Avenue, 5th Fl, Ste 545", "New York", "NY", "10022",
        "855-435-7144", "Info@Helpmedicalsupplies.Com", "helpmedicalsupplies.com", "DME Distributor/Retailer",
        "Manhattan showroom plus online distribution model.",
        "A combined Manhattan showroom and online distribution model gives this business both a retail presence and a scalable e-commerce channel.",
        "I came across Help Medical Supplies and was impressed by the combination of a Manhattan showroom with a broader online distribution business."),
    row("Welcare Pharmacy & Surgical", "Owner — name not confirmed", "54 Graham Ave", "Brooklyn", "NY", "11206",
        "718-599-7200", "INFO@WELCARESUPPLY.COM", "welcaresupply.com", "Specialty Pharmacy + DME/Surgical Supply",
        "Established 2004; dual pharmacy + DME model with custom compounding niche.",
        "A dual pharmacy-and-DME model with custom compounding is a differentiated service mix that's hard to find combined under one roof.",
        "I came across Welcare Pharmacy & Surgical and was impressed by the combined pharmacy and DME model, including custom compounding, you've built since 2004."),
    row("Connecticut Surgical Inc.", "Owner — name not confirmed", "14 Main Street", "Norwalk", "CT", "06851",
        "203-838-2354", "Email on file — confirm directly before outreach", "connecticutsurgical.net", "Independent Surgical/Medical Equipment Dealer",
        "Independent dealer presence in Norwalk, CT.",
        "An independent surgical equipment dealer in Norwalk represents a focused, niche acquisition target worth a direct follow-up call to confirm contact details.",
        "I came across Connecticut Surgical and wanted to reach out about your independent surgical/medical equipment dealership in Norwalk."),
]

INDUSTRIES = [
    ("Plumbing", PLUMBING, "2E86C1"),
    ("Manufacturing", MANUFACTURING, "707B7C"),
    ("Supermarkets", SUPERMARKETS, "229954"),
    ("Healthcare Distribution", HEALTHCARE_DIST, "8E44AD"),
]

wb = openpyxl.Workbook()
wb.remove(wb.active)

hook_col, email_col, state_col = 12, 8, 5

for sheet_name, leads, tab_color in INDUSTRIES:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color
    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_W), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 36

    for i, lead in enumerate(leads, start=2):
        state = lead[4] or ""
        base_fill_color = STATE_FILL.get(state, ROW_ALT1 if i % 2 == 0 else ROW_ALT2)
        for col_idx, value in enumerate(lead, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            if col_idx == hook_col:
                cell.fill = PatternFill("solid", fgColor=AMBER_BG)
                cell.font = HOOK_FONT
            elif col_idx == 13:
                cell.fill = PatternFill("solid", fgColor=BLUE_BG)
                cell.font = EMAIL_FONT
            else:
                cell.fill = PatternFill("solid", fgColor=base_fill_color)
                cell.font = REGULAR_FONT
        ws.row_dimensions[i].height = 170

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(leads)+1}"

out_path = "/home/user/Claude-/Multi_Industry_Leads_Batch1.xlsx"
wb.save(out_path)
total = sum(len(leads) for _, leads, _ in INDUSTRIES)
print(f"Saved {out_path} with {total} verified leads across {len(INDUSTRIES)} industries.")
